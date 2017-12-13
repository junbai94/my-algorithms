# encoding: utf-8
# -*- coding: utf-8 -*-
'''
Author: Jax Xie
Summary:
    This package consists of a number of functions that is used by the Risk and Analytics team.
    Do note that some functions are experimental and not in use while some functions has high reusability.

Organisation:
    1) Automan Interfacing:
        -These set of function interface with the web services of Automan.
        -Currently it has more functions than the official package from TAT but could be less
         stable as it uses post and get the web interface url.
        -The code allow us to do simple CRUD as well as mass loading of data into automan from excel and vice versa
        -Pricestore loading, RV model all uses these functions.
        -Function name usually has prefix of 'automan'
    2) Model and Statistics:
        -The RV model main algorithm is in this package as well. Function is 'get_relative_value_v2'. 'permutate_model' is used as well.
        -Statistics function such as Augmented Dickey-Fuller('adf_df'), orthogonal regression (orthoregress)
    3) Misc:
        -Other helper functions such as 'sort_tuple_replacement',_encrypt falls under this category



'''
import openpyxl
from openpyxl.worksheet import *
from openpyxl.utils import get_column_letter, column_index_from_string
import pickle
from bs4 import BeautifulSoup
import time
from datetime import date, timedelta, datetime
import requests
from requests_ntlm import HttpNtlmAuth
import shutil
import os
import copy
import logging
import urllib2
import codecs
import warnings
import numpy as np
import pandas as pd
from requests_negotiate_sspi import HttpNegotiateAuth
from scipy import stats

from shutil import copyfile
from urllib import quote

import math

from statsmodels.tsa.stattools import adfuller
import matplotlib.pylab as plt
from matplotlib.pylab import rcParams
rcParams['figure.figsize'] = 15, 6

import statsmodels.api as sm

from sklearn import linear_model

from scipy.odr import Model, Data, ODR
from scipy.stats.mstats import linregress

import sys
if sys.version_info[0] < 3:
    from StringIO import StringIO
else:
    from io import StringIO


import subprocess #for running vbscript
from dateutil.relativedelta import relativedelta #for adding dates
from datetime import timedelta, date

from urlparse import urlparse
from Crypto.Cipher import AES
import base64




#set up the proxy auth
proxyDict = None

#chrome browser header:
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}


_session_data = {
    "steelindex": {
        "email":"john_dai@cargill.com",
        "pass":"tsi",
        "btrLogin": "Login",
        "login_url" : "https://www.thesteelindex.com/en/restricted-area/"
    },
    "steelbb": {
        "username": "rick_zeng@cargill.com",
        "password": "tjjfyyae",
        "doLogin": "Login",
        "remember_me": 1,
        "login_url": "https://www.steelbb.com/login/"
    },
    "mysteel": {
        "my_username": "rioaussie",
        "my_password": "RIO2009",
        "site": "mysteel",
        "callback": "http://jiancai.mysteel.com/m/16/0729/11/F250E9B9835EC176.html",
        "login_url": "http://passport.mysteel.com/login.jsp"
    },
    "mysteel2": {
        "my_username": "rioaussie",
        "my_password": "RIO2009",
        "site": "mynengyuan",
        "callback": "http://coal.glinfo.com/p/16/1213/15/D055D470A59F813A.html",
        "login_url": "http://passport.mysteel.com/login.jsp"
    },
    # "mysteel": {
    #     "my_username": "cargill03",
    #     "my_password": "metals888",
    #     "site": "mysteel",
    #     "callback": "http://jiancai.mysteel.com/m/16/0729/11/F250E9B9835EC176.html",
    #     "login_url": "http://passport.mysteel.com/login.jsp"
    # }
}

automan_root = "http://ot-metals.automan.cargill.com"




def replace_text_csv_file(
        input_csv_path="C:\Users\j256377\Desktop\Metals\Minerva\sanitization\data\wechat\wechat_utf8.csv",
        output_csv_path="C:\Users\j256377\Desktop\Metals\Minerva\sanitization\data\wechat\wechat_sanitized.csv",
        replacement_tuples=[("old","new"),("xxx","yyy")],
        logger=None,
        display_log=True
    ):
    '''
    Function replace the text in the CSV file using the list of tuples from replacement_tuples.
    It then output the csv into output_csv_path
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    with codecs.open(input_csv_path, encoding='utf-8') as ifile:
        s = ifile.read()

    num_replacements = len(replacement_tuples)
    num_completed_replacements = 0
    for replacer in replacement_tuples:
        old=replacer[0]
        new=replacer[1]
        num_completed_replacements+=1
        if display_log:
            logger.info(u"=Replacing {} with {}".format(old,new))
            completed_percentage=float(num_completed_replacements)/float(num_replacements)*100
            logger.info(u"=={0:.2f}% completed. {1} out of {2} replacement made".format(completed_percentage,num_completed_replacements,num_replacements))

        s=case_insensitive_replace(old,new,s,by_word=True)




    with codecs.open(output_csv_path,'w',encoding='utf8') as ofile:
        ofile.write(s)



def _is_english(s):
    '''
    s is unicode. Check if it is an english string. Not if mixture of eng and chinese, will return false.
    '''
    try:
        s.decode('ascii')
    except:
        return False
    else:
        return True

def case_insensitive_replace(old=u"duck",new=u"chicken",haystack=u"dUk DuCk go duckling",by_word=True):
    '''
    Function does a case insentive replace.
    If by_word is true, it only replace word not character. e.g. if replacing "phil" with "test", 'phillipines' will not become 'testlipines'.
    By word only works for english 'old'. For chinese character we ignore
    e.g.
    _case_insensitive_replace(old=u"test",new=u"Not really",haystack=u"this is a test of 你好吗 ")
    #u'this is a Not really of \u4f60\u597d\u5417 '
    _case_insensitive_replace(old=u"你好",new=u"Not really",haystack=u"你好吗 this is a test of 你好吗")
    #u'Not really\u5417 this is a test of Not really\u5417'
    _case_insensitive_replace(old=u"你好吗",new=u"Not really",haystack=u"this is a test of 你好吗 ",by_word=False)
    #u'this is a test of Not really '
    '''
    s=haystack
    if not _is_english(old):#if not english word, i.e. chinese, we replace by_word
        by_word=False

    if by_word:
        pattern = re.compile(u"\\b{}\\b".format(old), re.IGNORECASE | re.UNICODE)
    else:
        pattern = re.compile(old, re.IGNORECASE)
    s=pattern.sub(new, s)

    return s


def demingregress_excel(X, Y,logger=None):
    '''
    Excel implemenation of deming.
    Return [Slope, Intercept,rPearson]
    '''
    if not logger:
        logger = logging.getLogger(__name__)
    try:
        meanX=[]
        meanY=[]
        meanX2=[]
        meanY2=[]
        meanXY=[]
        deltaX2=[]
        deltaY2=[]
        parameter=[]

        Nvalue=len(X)

        X=[float(i) for i in X]
        Y=[float(i) for i in Y]

        for j in range(0,Nvalue,2):

            if j<=Nvalue-2:
                meanX.append((X[j]+X[j+1])/2)
                meanY.append((Y[j]+Y[j+1])/2)
                deltaX2.append((X[j]-X[j+1])**2)
                deltaY2.append((Y[j]-Y[j+1])**2)
            else:break
        Npairs=int(Nvalue/2)
        for j in range(Npairs):
            meanX2.append(meanX[j]**2)
            meanY2.append(meanY[j]**2)
            meanXY.append(meanX[j]*meanY[j])

        sumX=sum(meanX)
        sumY=sum(meanY)
        sumX2=sum(meanX2)
        sumY2=sum(meanY2)
        sumXY=sum(meanXY)
        sumdeltaX2=sum(deltaX2)
        sumdeltaY2=sum(deltaY2)


        Xbar=sumX/Npairs
        Ybar=sumY/Npairs
        Sx2 = (Npairs * sumX2 - sumX**2) / (Npairs * (Npairs - 1))
        Sy2 = (Npairs * sumY2 - sumY**2) / (Npairs * (Npairs - 1))
        Sdx2 = sumdeltaX2 / float((2 * Npairs))
        Sdy2 = sumdeltaY2 / float((2 * Npairs))

        rPearson = (Npairs * sumXY - sumX * sumY)/math.sqrt((Npairs * sumX2 - sumX**2) * (Npairs * sumY2 - sumY**2))
        lamb = Sdx2 / Sdy2
        U = (Sy2 - Sx2 / lamb) / (2 * rPearson * math.sqrt(Sx2) * math.sqrt(Sy2))
        Slope = U + math.sqrt(U**2 + 1 / lamb)
        Intercept = Ybar - Slope * Xbar
        parameter=[Slope, Intercept, rPearson]

        return parameter
    except:
        e = sys.exc_info()[0]
        logger.warn("[__demingregress_excel]: Error in demingregress_excel. Returning 0 for all {}".format(e))
        return [0,0,0]






def demingregress(x, y,w_intercept=True,delta=1,print_stat=False):
    '''
    Deming_regression
    Delta refer to https://en.wikipedia.org/wiki/Deming_regression and http://clinchem.aaccjnls.org/content/clinchem/25/3/432.full.pdf
    x and y best in list. x and y must be same length
    '''
    if len(x) != len(y):
        raise ValueError("Length of x is not the sanem as length of y")
    n=float(len(x))
    mean_x = np.mean(x)
    mean_y = np.mean(y)
    s_xx = 1.0/(n-1)*np.sum([np.square(e-mean_x) for e in x])
    s_yy = 1.0/(n-1)*np.sum([np.square(d-mean_y) for d in y])
    s_xy = 1.0/(n-1)*np.sum([(f[0]-mean_x)*(f[1]-mean_y) for f in zip(x,y)])
    slope_wiki=(s_yy-delta*s_xx+np.sqrt(np.square(s_yy-delta*s_xx)+4*delta*np.square(s_xy)))/(2*s_xy)
    intercept_wiki = mean_y - slope_wiki * mean_x

    sum_delta_x2=np.sum([np.square(a-b) for a,b in zip(x,x[1:])])
    sum_delta_y2=np.sum([np.square(a-b) for a,b in zip(y,y[1:])])
    sdx2=sum_delta_x2/n
    sdy2=sum_delta_y2/n
    lmbd=sdx2/sdy2
    U=(s_yy-1.0/lmbd*s_xx)/(2*s_xy)
    slope=U + np.sqrt(np.square(U)+1.0/lmbd)
    intercept = mean_y - slope * mean_x

    if print_stat:
        print "mean_x: {}".format(mean_x)
        print "mean_y: {}".format(mean_y)
        print "s_xx: {}".format(s_xx)
        print "s_yy: {}".format(s_yy)
        print "s_xy: {}".format(s_xy)
        print "delta: {}".format(delta)
        print "slope_wiki: {}".format(slope_wiki)
        print "intercept_wiki: {}".format(intercept_wiki)
        print "lmbd: {}".format(lmbd)
        print "U: {}".format(U)
        print "slope: {}".format(slope)
        print "intercept: {}".format(intercept)

    return (slope,intercept)


def orthoregress(x, y,w_intercept=True):
    """Perform an Orthogonal Distance Regression on the given data,
    using the same interface as the standard scipy.stats.linregress function.
    Arguments:
    x: x data
    y: y data
    Returns:
    [m, c, nan, nan, nan]
    Uses standard ordinary least squares to estimate the starting parameters
    then uses the scipy.odr interface to the ODRPACK Fortran code to do the
    orthogonal distance calculations.
    """
    linreg = linregress(x, y)
    dat = Data(x, y)

    if w_intercept:
        mod = Model(linear_f_for_orthoregress)
        od = ODR(dat, mod, beta0=linreg[0:2])
    else:
        mod = Model(linear_f_for_orthoregress_no_intercept)
        od = ODR(dat, mod, beta0=linreg[0:1])

    out = od.run()

    return list(out.beta) + [np.nan, np.nan, np.nan]

def linear_f_for_orthoregress(p, x):
    """Basic linear regression 'model' for use with ODR"""
    return (p[0] * x) + p[1]

def linear_f_for_orthoregress_no_intercept(p, x):
    """Basic linear regression 'model' for use with ODR"""
    return (p[0] * x) #+ p[1]

def test_stationarity(timeseries):
    '''
    Function will plot charts and the results from Dickey-Fuller Test
    Taken from https://www.analyticsvidhya.com/blog/2016/02/time-series-forecasting-codes-python/
    '''
    #Determing rolling statistics
    rolmean = pd.rolling_mean(timeseries, window=12)
    rolstd = pd.rolling_std(timeseries, window=12)

    #Plot rolling statistics:
    orig = plt.plot(timeseries, color='blue',label='Original')
    mean = plt.plot(rolmean, color='red', label='Rolling Mean')
    std = plt.plot(rolstd, color='black', label = 'Rolling Std')
    plt.legend(loc='best')
    plt.title('Rolling Mean & Standard Deviation')
    plt.show(block=False)

    #Perform Dickey-Fuller test:
    print 'Results of Dickey-Fuller Test:'
    dftest = adfuller(timeseries, autolag='AIC')
    dfoutput = pd.Series(dftest[0:4], index=['Test Statistic','p-value','#Lags Used','Number of Observations Used'])
    for key,value in dftest[4].items():
        dfoutput['Critical Value (%s)'%key] = value
    print dfoutput

def adf_df(df,value_type=None,wide_table=True):
    '''
    Take a df and run augmented dickey fuller test. Each col of df is an asset and each row is ratio of asset 1 to asset 2..g. df:
            Date                           asset     PRICESTORE_DCE_CKCCNY_STORE_T3     PRICESTORE_DCE_IOECNY_STORE_T3  PRICESTORE_DCE_KEECNY_STORE_T3     PRICESTORE_SGX_HRC_STORE_T3  PRICESTORE_SHFE_HRCCNY_STORE_T3  PRICESTORE_SHFE_RBCNY_STORE_T3
0 2014-09-01         PRICESTORE_DCE_CKCCNY_STORE_T3                             NaN                                  0                             NaN                             NaN   0                          NaN                              NaN
    Optional:
        value_type: The type of data from the ADF test to be return. By default return everything (as tuple) as per http://statsmodels.sourceforge.net/0.5.0/generated/statsmodels.tsa.stattools.adfuller.html
        wide_table: if True, return a wide table with each asset as column. Else return two column of asset and asset2. Default true

    '''
    logger = logging.getLogger(__name__)

    col_date = "Date"
    col_asset = "asset"

    assets_list = df.columns.tolist()
    #remove date and assets
    try:
        assets_list.remove(col_date)
        assets_list.remove(col_asset)
    except:
        pass


    df_dict = {"asset":[]}
    return_dict = {}

    #prepare the dict
    for asset in assets_list:
        df_dict[asset]=[]

    #get the result
    for asset1 in assets_list:
        for asset2 in assets_list:
            #select the relevant "Date","asset","<asset2>"
            series = df[df[col_asset]==asset1]
            series = series[[col_date,asset2]]
            series = series.sort_values(col_date,ascending=True)
            series = series[asset2]

            if asset1 not in return_dict:
                return_dict[asset1] = {}

            return_value = None

            if asset1 != asset2:
                logger.info("-- {} vs {} --".format(asset1,asset2))
                adf_result = sm.tsa.stattools.adfuller(series.dropna()) #http://statsmodels.sourceforge.net/0.5.0/generated/statsmodels.tsa.stattools.adfuller.html
                if value_type=="p":
                    return_value = adf_result[1]
                else:
                    return_value = adf_result

                return_dict[asset1][asset2] = return_value

            else:
                return_dict[asset1][asset2]=None



            if asset2 == assets_list[0]: #for first element... we add a row to asset
                df_dict["asset"].append(asset1)
            df_dict[asset2].append(return_value)



    df = pd.DataFrame(df_dict)[['asset']+assets_list]

    if not wide_table:
        df = pd.melt(df, id_vars=['asset'], value_vars=assets_list,var_name="asset2")

    return {"dict":return_dict,"df":df}


def df_corr(df,args):
    correlation_coeff = df.corr(method=args['method']).iloc[0,1]

    return correlation_coeff

def _rolling_apply_w_date(
            df=pd.DataFrame({
                            "Date": pd.Series(pd.date_range('1/1/2016', periods=366, freq='D')),
                            "price1": np.random.randn(366)*10,
                            "price2": np.random.randn(366)*10,
                            })
            ,window=5,func=df_corr,args={"method":'pearson'}
            ):
    '''
    Function takes a df with 3 columns only (including Date), and apply func across them.
    func prototype is func(dataframe,args). func must return single value
    return the same df but with additional column of 'calculation'

    e.g. _rolling_apply_w_date(df2,10,df_corr,args={"method":"pearson"})

    e.g.  df
         Date  PRICESTORE_SHFE HRC ACTIVE1_ACTIVE1  PRICESTORE_SHFE HRC ACTIVE1_ACTIVE1
2015-10-07                               1862.0                              1862.0
2015-10-08                               1863.0                              1863.0

    e.g. return result:
         Date  PRICESTORE_SHFE HRC ACTIVE1_ACTIVE1  PRICESTORE_SHFE HRC ACTIVE1_ACTIVE1  calculation
2015-10-07                               1862.0                              1862.0             NaN
2015-10-08                               1863.0                              1863.0             0.98


    '''
    def _window_func(ii,df,kwargs):
        x_df = df.iloc[map(int, ii)]
        # result =x_df[[kwargs['col1_name'],kwargs['col2_name']]].corr(method='pearson').iloc[0,1]
        result =func(x_df[[kwargs['col1_name'],kwargs['col2_name']]],args=kwargs)
        return result

    df_index=pd.DataFrame({"ii":range(len(df))})
    col_names = df.columns.tolist()
    col_names.remove('Date')

    args.update({"col1_name":col_names[0],"col2_name":col_names[1]})

    result = df_index.rolling(window=window)\
        .apply(func=lambda x: _window_func(x,df,args))
    calc_df = pd.DataFrame(result)
    calc_df.columns=['calculation']

    # return calc_df
    #join with data frame by index
    df=df.reset_index()
    calc_df=calc_df.reset_index()
    result_df = pd.concat([df,calc_df],axis=1,join_axes=[df.index])


    result_df.drop('index', axis=1, inplace=True)

    return result_df

def get_correlation(
                        df=pd.DataFrame({
                                    "Date": pd.Series(pd.date_range('1/1/2016', periods=366, freq='D')),
                                    "price1": np.random.randn(366)*10,
                                    "price2": np.random.randn(366)*10,
                                    }),
                        col_date="Date",
                        start_date = date(2016, 1, 1),
                        end_date = date(2016, 12, 31),
                        dropna=False, #drop row with na....
                        mode = "all",
                        get_percentile = False, #no effect
                        get_daily_rank = False, #if you want a column of rank. Note agg_row_func must not be None for this to work
                        window_num_days=10,#rolling windows for corr, set to None if don't want moving windows
                        logger=None,
                        wide_table = False, #no effect
                        tall_table_col_name = [] # no effect

    ):


    '''
    Get the correlation of two time series

    Optional:
        mode: "all" return pearson,kendall and spearman coefficient
              "pearson" return just pearson coeff
              "kendall" return just kendall coeff
              "spearman" return just spearman coeff
    '''

    logger = logging.getLogger(__name__)
    if not logger:
        logger = logging.getLogger(__name__)
    col_date = df.columns[0]
    col_name_1 = df.columns[1]
    col_name_2 = df.columns[2]

    logger.info("-- {} vs {} --".format(col_name_1,col_name_2))

    col_calculation = mode


    if mode=="all":
        tall_table_col_name=["pearson","kendall","spearman","ts1","ts2"]
    elif mode == "pearson":
        tall_table_col_name=["pearson","ts1","ts2"]
    elif mode == "kendall":
        tall_table_col_name=["kendall","ts1","ts2"]
    elif mode == "spearman":
        tall_table_col_name=["spearman","ts1","ts2"]

    #for same column...
    if col_name_1 == col_name_2:
        if wide_table:
            null_values = [np.NaN]*df[col_name_2].shape[0]
            col_name_tmp = "col_name_tmp"
            df[col_name_tmp]=null_values
            df = df[[col_date,col_name_tmp]] #return date,col_name_tmp which is all NaN
            df.columns = [col_date,col_name_2]
            return df
        else:
            null_values = [np.NaN]*df[col_name_2].shape[0]
            df['asset2']=col_name_2
            col_interest = tall_table_col_name
            for col in col_interest:
                df[col]=np.NaN
            df = df[[col_date,'asset2']+col_interest]
            return df

    #1) remove na data and sort by date...
    logger.info(start_date)
    df = df[(df[col_date] >= start_date) & (df[col_date] < end_date)]
    if dropna:
        df = df.dropna()
    df = df.sort_values(col_date,ascending=True)

    #2) Do the calculation
    #Get the correlation
    if mode=="all":
        #no moving window correlation
        if not window_num_days:
            pearson = df.corr(method='pearson').iloc[0,1]
            kendall = df.corr(method='kendall').iloc[0,1]
            spearman = df.corr(method='spearman').iloc[0,1]
        else:
            pearson = _rolling_apply_w_date(df,window_num_days,df_corr,args={"method":"pearson"})['calculation']
            kendall = _rolling_apply_w_date(df,window_num_days,df_corr,args={"method":"kendall"})['calculation']
            spearman = _rolling_apply_w_date(df,window_num_days,df_corr,args={"method":"spearman"})['calculation']
        df['pearson']=pearson.tolist() #without to list, pandas may try a join by index given that pearson is a pandas series
        df['kendall']=kendall.tolist()
        df['spearman']=spearman.tolist()
        df['ts1']=df[col_name_1]
        df['ts2']=df[col_name_2]
    elif mode == "pearson":
        pearson = df.corr(method='pearson').iloc[0,1]
        df[col_calculation]=pearson
    elif mode == "kendall":
        kendall = df.corr(method='kendall').iloc[0,1]
        df[col_calculation]=kendall
    elif mode == "spearman":
        spearman = df.corr(method='spearman').iloc[0,1]
        df[col_calculation]=spearman


    df['asset2']=col_name_2
    col_interest = [col_date,'asset2']+tall_table_col_name
    df = df[col_interest]


    return df

def get_relative_value_v2(
                        df=pd.DataFrame({
                                    "Date": pd.Series(pd.date_range('1/1/2016', periods=366, freq='D')),
                                    "price1": np.random.randn(366)*10, #these are normalized price. e.g. before converting to say USD from CNY
                                    "price2": np.random.randn(366)*10, #these are normalized price

                                    "price1_denormalize": np.random.randn(366)*10, #these are denormalized original price
                                    "price2_denormalize": np.random.randn(366)*10,
                                    }),
                        # window_num_days=90,
                        col_date="Date",
                        start_date = date(2016, 1, 1),
                        end_date = date(2016, 12, 31),
                        dropna=False, #drop row with na....
                        mode = "regression",
                        get_percentile = False,
                        get_daily_rank = False, #if you want a column of rank. Note agg_row_func must not be None for this to work
                        window_num_days=90,#optional
                        wide_table = True, #if set to False,
                        logger=None,
                        tall_table_col_name = [] # if wide_table is set to none, this has to be set. e.g. ["slope","intercept","regressed_cal","asset1_diff","asset2_diff","asset1_price","asset2_price"]

    ):


    '''
    Get difference between the prices by comparing with previous value. Do so for both prices.
    Regress these 2 difference based on window period and get the slope.
    Use the slope to compute price1-slope*price2

    Optional:
        mode: "regression" return the calculated regression. Y is price1 and X is price2
              "regression_all" return everything, i.e. slope,interception etc. Only avaliable in tall table mode. i.e. wide_table==False
              "difference" return the calculated difference price1 - price2
              "ratio" return the price1/price2
              "percentile" to return percentile
    '''
    if not logger:
        logger = logging.getLogger(__name__)
    col_date = df.columns[0]
    col_name_1 = df.columns[1]
    col_name_2 = df.columns[2]
    if len(df.columns)>=5: #use the denormalize df if it exists
        col_name_1_denormalize = df.columns[3]
        col_name_2_denormalize = df.columns[4]
    else: #else don't exists we just use the only price available
        col_name_1_denormalize = df.columns[1]
        col_name_2_denormalize = df.columns[2]

    logger.info("-- {} vs {} --".format(col_name_1,col_name_2))


    col_name_1_diff = "{}_diff".format(col_name_1)
    col_name_2_diff = "{}_diff".format(col_name_2)
    col_calculation = mode


    #for same column...
    if col_name_1 == col_name_2:
        if wide_table:
            null_values = [np.NaN]*df[col_name_2].shape[0]
            col_name_tmp = "col_name_tmp"
            df[col_name_tmp]=null_values
            df = df[[col_date,col_name_tmp]] #return date,col_name_tmp which is all NaN
            df.columns = [col_date,col_name_2]
            return df
        else:
            null_values = [np.NaN]*df[col_name_2].shape[0]
            df['asset2']=col_name_2
            col_interest = [col_calculation]+tall_table_col_name
            if get_percentile: #if get percentile the calculation is percentile
                col_interest.append('percentile')
                if mode=="regression_all": #for regression_all we get percentile of ratio too.
                    col_interest.append('percentile_ratio')

            for col in col_interest:
                df[col]=np.NaN

            df = df[[col_date,'asset2']+col_interest]



            return df


    #1) remove na data and sort by date...
    # logger.info("==MAX DATE: {}==".format(max(df[col_date])))
    df = df[(df[col_date] >= start_date) & (df[col_date] < end_date)]
    # logger.info("==MAX DATE AFTER FILTER: {}==".format(max(df[col_date])))

    if dropna:
        df = df.dropna()
    df = df.sort_values(col_date,ascending=True)

    df[col_name_1_diff]=df[col_name_1]-df[col_name_1].shift(1)
    df[col_name_2_diff]=df[col_name_2]-df[col_name_2].shift(1)

    df = df.dropna() #this will remove the earliest date which has no diff.



    #for empty df...
    if df.empty:
        logger.warn("===[] EMPTY Data detected after finding difference for {} vs {}===".format(col_name_1,col_name_2))

        null_values = [np.NaN]*df[col_name_2].shape[0]
        if wide_table:
            col_name_tmp = "col_name_tmp"
            df[col_name_tmp]=null_values
            df = df[[col_date,col_name_tmp]] #return date,col_name_tmp which is all NaN
            df.columns = [col_date,col_name_2]
            return df
        else:
            null_values = [np.NaN]*df[col_name_2].shape[0]
            df['asset2']=col_name_2
            col_interest = [col_calculation]+tall_table_col_name
            if get_percentile:
                col_interest.append('percentile')
                if mode=="regression_all": #for regression_all we get percentile of ratio too.
                    col_interest.append('percentile_ratio')
            for col in col_interest:
                df[col]=np.NaN
            df = df[[col_date,'asset2']+col_interest]

            return df


    if mode == "regression" or mode == "residual" or mode == "regression_all":
        #1. OLS linear regresion
            #get the slope
        # length = df[col_name_1_diff].shape[0]
        # x = df[col_name_2_diff].reshape(length,1)
        # y = df[col_name_1_diff].reshape(length,1)
        #
        # regr = linear_model.LinearRegression()
        # regr.fit(x,y)
        # slope = regr.coef_[0][0]
        # intercept = regr.intercept_[0]

        #2. Deming (aka orthogonal regression) https://en.wikipedia.org/wiki/Deming_regression
        x = df[col_name_2_diff].dropna()
        y = df[col_name_1_diff].dropna()
        # para = orthoregress(x,y,w_intercept=True)
        # para = demingregress(x,y)
        para = demingregress_excel(x,y,logger)

        slope = para[0]
        intercept = para[1]

        # print "len(x) {}, len(y) {}, para {}".format(len(x),len(y),para)
        # print "x.head(10) {} y.head(10) {}".format(x.head(10),y.head(10))

        # x.to_csv("x{}.csv".format(col_name_2),header=["x"],index=False)
        # y.to_csv("y{}.csv".format(col_name_1),header=["y"],index=False)

        if np.isnan(intercept): #for Nan intercept (in case where no intercept), we set to zero
            intercept = 0


        # print df[col_name_1].head(1)
        # print df[col_name_2].head(1)
        # print slope
        # print intercept

        if mode == "regression":
            #Get the calculated value (be it regressed or ratio or diff...)
            df[col_calculation]=df[col_name_1]-slope*df[col_name_2]
        elif mode=="residual":
            # df[col_calculation]=y-regr.predict(x)
            df[col_calculation]=y-(x*slope+intercept)
        elif mode=="regression_all":
            if slope > 0:
                df[col_calculation]=df[col_name_1]-slope*df[col_name_2]
            else:
                df[col_calculation]=df[col_name_1]+slope*df[col_name_2]
            df_regress_all=pd.DataFrame({"slope":slope,
                                        "intercept":intercept,
                                        "regressed_cal":df[col_calculation],#df[col_name_1]-slope*df[col_name_2],
                                        "asset1_diff":df[col_name_1_diff],
                                        "asset2_diff":df[col_name_2_diff],
                                        "asset1_price":df[col_name_1],
                                        "asset2_price":df[col_name_2],
                                        "asset1_price_denormalize":df[col_name_1_denormalize],
                                        "asset2_price_denormalize":df[col_name_2_denormalize],
                                        "asset1_price_asset2_price_ratio":df[col_name_1]/df[col_name_2],
                                        })

            # df_regress_all['combined_all']=df_regress_all.apply(lambda row: {
            #                         "slope":row["slope"],
            #                         "intercept":row["intercept"],
            #                         "regressed_cal":row["regressed_cal"],
            #                         "asset1_diff":row["asset1_diff"],
            #                         "asset2_diff":row["asset2_diff"],
            #                         "asset1":row["asset1"],
            #                         "asset2":row["asset2"],
            #                         },axis=1)
            # df[col_calculation+"_all"]=df_regress_all['combined_all']

            df['slope']=df_regress_all['slope']
            df['intercept']=df_regress_all['intercept']
            df['regressed_cal']=df_regress_all['regressed_cal']
            df['asset1_diff']=df_regress_all['asset1_diff']
            df['asset2_diff']=df_regress_all['asset2_diff']
            df['asset1_price']=df_regress_all['asset1_price']
            df['asset2_price']=df_regress_all['asset2_price']
            df['asset1_price_denormalize']=df_regress_all['asset1_price_denormalize']
            df['asset2_price_denormalize']=df_regress_all['asset2_price_denormalize']
            df['asset1_price_asset2_price_ratio']=df_regress_all['asset1_price_asset2_price_ratio']

    elif mode == "difference":
        df[col_calculation]=df[col_name_1]-df[col_name_2]
    elif mode == "ratio":
        df[col_calculation]=df[col_name_1]/df[col_name_2]
        #remove the infinity value
        df = df.replace([np.inf, -np.inf], np.nan)
        df = df.dropna()



    values = [] #stored either the regressed value or percentile
    values2 = [] #stored the percentile for ratio in case of mode=="regression_all"

    for index, row in df.iterrows():
        if(
            not get_percentile
        ):
            values.append(row[col_calculation])
        else:
            percentiles_row = stats.percentileofscore(df[col_calculation],row[col_calculation],kind='rank')
            values.append(percentiles_row)

            if mode == "regression_all": #for regression all we calculate percentile for the ratio too
                percentiles_row2 = stats.percentileofscore(df["asset1_price_asset2_price_ratio"],row["asset1_price_asset2_price_ratio"],kind='rank')
                values2.append(percentiles_row2)
            # print col_name_1,col_name_2
            # if col_name_1=="Coke T1" and col_name_2=="Coke T2":
            #     print "{} vs {} at {} Percentile: {}".format(col_name_1,col_name_2,row['Date'],percentiles_row)
            #     logger.info("{} vs {} at {} Percentile: {}".format(col_name_1,col_name_2,row['Date'],percentiles_row))


    df[col_name_2]=values

    # if mode=="regression_all":
    #     # print df.columns
    #
    #     #!For some unknown reason we need to convert dictionary to string so we can store...
    #     df[col_name_2] = df.apply(lambda row: str({
    #                             "value":row[col_name_2],
    #                             "slope":row['slope'],
    #                             "intercept":row["intercept"],
    #                             "regressed_cal":row["regressed_cal"],
    #                             "asset1_diff":row["asset1_diff"],
    #                             "asset2_diff":row["asset2_diff"],
    #                             "asset1_price":row["asset1_price"],
    #                             "asset2_price":row["asset2_price"]
    #                             }),axis=1)
    #     # dict_result = (df.apply(to_dict_regression,axis=1))

    if wide_table:
        df = df[[col_date,col_name_2]] #return date,col_name2
    else:
        #NEED COLUMN OF asset and asset2
        df['asset2']=col_name_2
        col_interest = [col_date,'asset2',col_calculation]+tall_table_col_name
        if get_percentile: #if get percentile the calculation is percentile
            df['percentile']=values
            col_interest.append('percentile')
            if mode=="regression_all":
                df['percentile_ratio']=values2
                col_interest.append('percentile_ratio')







        df = df[col_interest]


    return df



def get_ratio_value(
                        df=pd.DataFrame({
                                    "Date": pd.Series(pd.date_range('1/1/2016', periods=366, freq='D')),
                                    "price1": np.random.randn(366)*10,
                                    "price2": np.random.randn(366)*10,
                                    }),
                        window_num_days=90,
                        col_date="Date",
                        start_date = date(2016, 1, 1),
                        end_date = date(2016, 12, 31),
                        dropna=False, #drop row with na....
                        mode=None
    ):
    '''
    Get a ratio of price1/price2.
    df must contain only 2 prices and date.e.g.
                   Date  D_CHINAPRICES_SHANGHAI_REBAR  D_CHINAPRICES_TANGSHAN_BILLET  \
    0    2008-01-01                        4580.0                         4300.0
    1    2008-01-02                        4580.0                         4300.0
    2    2008-01-03                        4580.0                         4300.0
    3    2008-01-04                        4580.0                         4300.0
    4    2008-01-07                        4580.0                         4300.0
    Note df cannot have na value, if not it will be drop...
    Will return
                   Date    D_CHINAPRICES_TANGSHAN_BILLET
    0    2008-01-01                       1.1
    1    2008-01-02                       0.9
    '''
    logger = logging.getLogger(__name__)
    col_date = df.columns[0]
    col_name_1 = df.columns[1]
    col_name_2 = df.columns[2]

    df = df[(df[col_date] >= start_date) & (df[col_date] < end_date)]
    if dropna:
        df = df.dropna()


    # df.loc[:,'ratio'] = df[col_name_1]/df[col_name_2]
    if col_name_1==col_name_2:
        df['ratio'] = pd.Series(np.nan,index=range(0,df.shape[0]))
    else:
        df['ratio'] = df[col_name_1]/df[col_name_2]



    result_df = df[[col_date,"ratio"]]
    result_df.columns = [col_date,col_name_2]

    return result_df



def get_relative_value(
                        df=pd.DataFrame({
                                    "Date": pd.Series(pd.date_range('1/1/2016', periods=366, freq='D')),
                                    "price1": np.random.randn(366)*10,
                                    "price2": np.random.randn(366)*10,
                                    }),
                        window_num_days=90,
                        col_date="Date",
                        start_date = date(2016, 1, 1),
                        end_date = date(2016, 12, 31),
                        dropna=False, #drop row with na....
                        mode=None
    ):
    '''
    Add a column to the df relative_value which shows the percentile it is at compared to the windo period
    required:
    df must contain only 2 prices and date.e.g.
                   Date  D_CHINAPRICES_SHANGHAI_REBAR  D_CHINAPRICES_TANGSHAN_BILLET  \
    0    2008-01-01                        4580.0                         4300.0
    1    2008-01-02                        4580.0                         4300.0
    2    2008-01-03                        4580.0                         4300.0
    3    2008-01-04                        4580.0                         4300.0
    4    2008-01-07                        4580.0                         4300.0
    Note df cannot have na value, if not it will be drop...
    '''
    logger = logging.getLogger(__name__)
    col_name_1 = df.columns[1]
    col_name_2 = df.columns[2]

    col_name_tmp_1 ="price1"
    col_name_tmp_2 ="price2"

    col_date = 'Date'
    col_ratio = ('{}_{}_{}').format(col_name_1,col_name_2,'ratio')
    col_percentiles = '{}_{}_percentiles'.format(col_name_1,col_name_2)

    #1) remove na data and sort by date...
    df = df[(df[col_date] >= start_date) & (df[col_date] < end_date)]
    if dropna:
        df = df.dropna()
    df = df.sort_values(col_date,ascending=True)
    df.index = range(0,len(df))

    df.columns = [col_date,col_name_tmp_1,col_name_tmp_2]
    df[col_ratio]=df[col_name_tmp_1]/df[col_name_tmp_2]

    percentiles = []

    for index, row in df.iterrows():
        start_row_id = index-window_num_days
        end_row_id = index
        if (
                (start_row_id < 0 or end_row_id >= df.shape[0])
                or
                (col_name_1==col_name_2)
            ): #invalid window periods or if comparing same prices...
            percentiles.append(None)
        else:
            percentiles_row = stats.percentileofscore(df[start_row_id:end_row_id][col_ratio],row[col_ratio],kind='rank')
            percentiles.append(percentiles_row)

    df[col_name_2]=percentiles

    df = df[[col_date,col_name_2]] #return date,col_name2

    return df


def sort_tuple_replacement(
    tuple_list=[("test","real"),("testing","really")],
    order="desc" #or "acs"
):
    '''
    Function sort a list of tuple. Each tuple is a double. The first one is the old text to be replace, and the second one is
    '''
    if order=="asc":
        reverse = False
    else:
        reverse = True

    tuple_list.sort(key=lambda t: len(t[0]), reverse=reverse)
    return tuple_list


def _get_human_readable_time_diff(start_time,end_time=datetime.now()):
    rd = relativedelta(end_time, start_time)
    output_str=""
    if rd.years > 0:
        output_str = output_str+"%d yrs" % rd.years
    if rd.months > 0:
        output_str = output_str+" %d mth" % rd.months
    if rd.days > 0:
        output_str = output_str+" %d days" % rd.days
    if rd.hours > 0:
        output_str = output_str+" %d hrs" % rd.hours
    if rd.minutes > 0:
        output_str = output_str+" %d mins" % rd.minutes
    if rd.seconds > 0:
        output_str = output_str+" %d secs" % rd.seconds
    if rd.microseconds > 0:
        output_str = output_str+" %d ms" % (rd.microseconds/1000)
    return output_str.strip()


def permutate_model(
    df,
    func,
    agg_row_func=None,
    col_date="Date",
    window_num_days=90,
    start_date = date(2012, 1, 1),
    end_date = date(2016, 8, 8),
    output_path=None,
    mode=None, #the mode to passed to the function 'func'. if mode is regression_all, then get_percentile will be true
    dropna = True,
    get_percentile=False,
    get_daily_rank=False,
    wide_table = True, #if set to False, will stack up other columns into asset2
    logger=None,
    pickle_output_path=None,#the path to store the wide table in pickled form, e.g. "c:\\tmp\\rv_pickle.p"
    pickle_input_path=None, #if this is set, then the loop for all assets wouldn't run. final_df will be loaded from this path
    tall_table_col_name=[],
    denormalize_mode=False, #if set to true, code will look for '[column_name]_denormalize'. This is the figure before denormalizing and will not be shown in calculation. But will be spilt out at the final result..
    ):
    '''
    Read df consisting of date and prices. e.g.
         Date          PRICESTORE_DCE_CKCCNY_STORE_T3  PRICESTORE_DCE_IOECNY_STORE_T3     PRICESTORE_DCE_KEECNY_STORE_T3
    518  25-Aug-2016                            830.0                           459.0                             1254.0
    519  26-Aug-2016                            830.0                           454.5                             1254.0
    520  27-Aug-2016                            828.0                           449.5                             1254.0
    521  28-Aug-2016                            828.0                           440.0                             1254.0
    522  29-Aug-2016                            828.0                           434.0                             1254.0
    It will apply func on all permutation of the prices and return a matrix with all possible dates and combination
    Optional:
        agg_row_func: can be "mean","max","min". aggregate across row and add a new column
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    sys_start_time = datetime.now()

    if mode=="regression_all":
        get_percentile=True;

    #Get list of prices...
    list_assets = df.columns[1:].tolist()
    final_df = None

    if denormalize_mode:
        list_assets_remove_denormalize = [ x for x in list_assets if "denormalize" not in x ]
        total_comparision = len(list_assets_remove_denormalize)*len(list_assets_remove_denormalize)
    else:
        total_comparision = len(list_assets)*len(list_assets)
    current_total_comparision=0



    if pickle_input_path==None: #if not pickled df to use...
        for first_asset in list_assets:
            first_assets_df = None
            if denormalize_mode and ("_denormalize" in first_asset):
                continue #move on to the next in loop, don't need to work on denormalize one
            for second_asset in list_assets:
                if denormalize_mode and ("_denormalize" in second_asset):
                    continue #move on to the next in loop, don't need to work on denormalize one

                log_msg = "Comparing {} with {}".format(first_asset,second_asset)
                logging.debug(log_msg)

                if denormalize_mode: #for denormalize_mode we passed the denormalized column as well
                    col_to_pass = [col_date,first_asset,second_asset,"{}_denormalize".format(first_asset),"{}_denormalize".format(second_asset)]
                else:
                    col_to_pass = [col_date,first_asset,second_asset]

                result_df = func(df[col_to_pass],
                                               col_date=col_date,
                                               window_num_days=window_num_days,
                                               start_date = start_date,
                                               end_date = end_date,
                                               mode = mode,
                                               dropna = dropna,
                                               get_percentile=get_percentile,
                                               wide_table=wide_table,
                                               tall_table_col_name=tall_table_col_name,
                                               logger=logger
                                               )


                if first_assets_df is None: #if no first_assets_df...
                    first_assets_df = result_df
                else:
                    if wide_table:
                        first_assets_df = first_assets_df.merge(result_df,how='outer',on='Date')
                    else: #not wide table we append
                        first_assets_df = first_assets_df.append(result_df)


                current_total_comparision=current_total_comparision+1
                percentage_completed = float(current_total_comparision)/float(total_comparision)*100.0
                logger.info("===={:.2f}% completed [Total Comparison Completed: {}, Total Comparison Needed: {} ] [Elapsed time: {}]====".format(percentage_completed,current_total_comparision,total_comparision,_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))

            col_names = first_assets_df.columns.tolist()
            first_assets_df['asset']=first_asset
            #rearrage the columns
            col_names.insert(1,'asset')
            first_assets_df = first_assets_df[col_names]


            if final_df is None: #if no first_assets_df...
                final_df = first_assets_df
            else:
                # print (final_df.columns)
                # print (first_assets_df.columns)
                final_df = pd.concat([final_df,first_assets_df])


            final_df = final_df[col_names]
            # logger.info("{} cols are {}".format(first_asset,col_names))
            # final_df.to_csv("c:/tmp_data/rv/test2/{}.csv".format(first_asset),index=False)

    else:
        final_df = pd.read_pickle(pickle_input_path)

    if pickle_output_path!=None and pickle_input_path==None: #save the pickle
        final_df.to_pickle(pickle_output_path)
    #calculate the aggregate across row
    if agg_row_func:
        agg_col = agg_row_func
        if agg_row_func == "mean":
            final_df[agg_col] = final_df[list_assets].mean(axis=1)
        elif agg_row_func == "max":
            final_df[agg_col] = final_df[list_assets].max(axis=1)
        elif agg_row_func == "min":
            final_df[agg_col] = final_df[list_assets].min(axis=1)
        else:
            pass

        #if we want to get the rank
        if get_daily_rank: #default get_daily_rank=False
            final_df=final_df.reset_index(drop=True)
            final_df['rank']=final_df.sort_values(col_date).groupby(col_date)[agg_col].rank(ascending=False)

    #check whether need to melt wide table
    # if not wide_table:
    #     logger.info("========================================")
    #     logger.info("====Melting DF... [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     final_df = pd.melt(final_df, id_vars=[col_date,'asset'], value_vars=list_assets,var_name="asset2")
    #     logger.info("====Melted DF! [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     logger.info("========================================")
    #
    # if mode=="regression_all":
    #     logger.info("========================================")
    #     logger.info("====Breaking down dict DF... [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     import ast
    #     import math
    #     def literal_dict_regression(x):
    #         if isinstance(x,float) and math.isnan(x):
    #             return {"value":np.nan,"intercept":np.nan,"regressed_cal":np.nan,"asset1_diff":np.nan,"asset2_diff":np.nan,"asset1_price":np.nan,"asset2_price":np.nan}
    #         else:
    #             try:
    #                 return ast.literal_eval(x)
    #             except:
    #                 print x
    #                 print x.__class__
    #                 raise
    #
    #     logger.info("========lambda apply [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     final_df_split_dict=final_df['value'].apply(lambda x: pd.Series(literal_dict_regression(x)))
    #     logger.info("========join [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     final_df = final_df[['Date','asset','asset2']].join(final_df_split_dict)
    #     logger.info("========rename [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     final_df.rename(columns={"value":"percentile"},inplace=True)
    #     logger.info("====Broke down dict DF! [Elapsed time: {}]====".format(_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
    #     logger.info("========================================")

    if output_path:
        logger.info("========================================")
        logger.info("====Writing to file {} [Elapsed time: {}]====".format(output_path,_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
        final_df.to_csv(output_path,index =False)
        logger.info("====Completed Writing to file {}! [Elapsed time: {}]====".format(output_path,_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now())))
        logger.info("========================================")

    return final_df



def generate_price1_2_parameters_frm_tags(
        tag="china_metal_prices",
        parameters_name="Price1"
    ):
    '''

    '''

    df = automan_read("china_metal_prices")

    price_list = df.columns.tolist()[1:] #ignore first column of datetime

    def remove_header(ts_name):
        ts_name = ts_name.replace("D_MYSTEEL_CHINA_PRICES_","")
        ts_name = ts_name.replace("D_MYSTEEL_","")
        ts_name = ts_name.replace("D_SHFE_","")
        ts_name = ts_name.replace("D_TSI_","")

        return ts_name

    def tableaurize(ts_name):
        ts_name = ts_name.replace("_"," ").title()
        return ts_name

    #1. Print the parameters to be copied to tableau
    for price in price_list:
        print "{},{}".format(tableaurize(price),remove_header(price))

    #2. Print Script...
    idx=0
    for price in price_list:
        if idx == 0:
            print "IF [{}]==\"{}\" THEN".format(parameters_name,tableaurize(price))
        else:
            print "ELSEIF [{}]==\"{}\" THEN".format(parameters_name,tableaurize(price))

        print " [{}]".format(tableaurize(price))

        if idx == len(price_list)-1:
            print "ELSE null"
            print "END"
        idx+=1




def _automan_tag_friendly(tag):
    '''
    Function turn a string and make it friendly for automan.
    Remove space and change all to small letters.
    Replace special characters too:
        "/" with "_per_"
        ":" with ""
        " " with "_"
    '''

    return str(tag).lower().replace(" ","_").replace("/","_per_")\
                           .replace(":","_")


def automan_update_ts_with_excel_col_as_tag(
        input_excel_path="C:\Users\j256377\Desktop\Metals\Automan Project\data dictionary\\tagging\\2017_01_09_data_dictionary_w_dean_tag.xlsx",
        name_col="A",
        start_col="G",
        end_col="M",
        logger=None
    ):
    '''
    Function read an excel file which contains a single sheet of:
        1) A column of time series (TS) name of the TS in automan specified by parameter name_col.
        2) Columns of attributes. First row must be header name. Specified by start_col and end_col. Include start_col and exclude end_col.
    It will add the tags to the time series in 'header:value'
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    #1. Read the excel file
    df = pd.read_excel(input_excel_path)

    #2. Loop through each row
    name_col_idx=column_index_from_string(name_col)-1
    start_col_idx=column_index_from_string(start_col)-1
    end_col_idx=column_index_from_string(end_col)-1
    cols = df.columns
    attributes = cols[start_col_idx:end_col_idx]

    for index, row in df.iterrows(): #for each time series
        tags=""
        ts_name = row[cols[name_col_idx]]
        #go through each attributes
        for attribute in attributes:
            value=row[attribute]
            if not pd.isnull(value): #only if not null then we add
                tags+="{}:{} ".format(_automan_tag_friendly(attribute),_automan_tag_friendly(value))

        tags = tags.strip()
        update_log = "updating ts_name:'{}' with \ntags '{}'...".format(ts_name,tags)
        logger.info("================{}======================".format(update_log))

        #update the tag for each time series
        automan_update_ts(name=ts_name,tags=tags)

        update_log = "updated ts_name:'{}' with \ntags '{}'!".format(ts_name,tags)
        logger.info("=================={}====================".format(update_log))

    pass


def automan_delete_by_tag(tags,logger=None):
    if not logger:
        logger = logging.getLogger(__name__)
    #Get the list of tags.
    df = automan_read(tags=tags,start_date=date(2117,1,1),convert_to_date=False)

    list_ts_name = df.columns.tolist()
    list_ts_name.remove('Date')

    num_ts = len(list_ts_name)
    idx=0
    for ts in list_ts_name:
        logger.info("{})Deleting {}.\n [{} of {} ts to be deleted]".format(str(idx+1),ts,str(idx+1),str(num_ts)))
        automan_delete_ts(ts,logger=None)
        logger.info("--Deleted! {}.\n [{} of {} ts deleted]".format(ts,str(idx+1),str(num_ts)))
        idx+=1

def automan_delete_ts(name,logger=None):
    '''
    Delete time series.
    '''
    if not logger:
        logger = logging.getLogger(__name__)
    #1. check whether time series exists
    if not automan_exists_ts(name):
        raise ValueError('Missing time series name "{}" in automan'.format())

    #2. take control ts
    automan_update_ts(name,take_control=True)

    #3. delete ts
    automan_update_ts(name,delete=True)




def automan_update_ts(name,take_control=False,delete=False,tags=None,new_name=None,description=None,overwrite_tags=False,overwrite_description=False,logger=None):
    '''
    Update the time series.
    Required:
        name: name of the series
    Optional
        tags: e.g. "nymex year:2016". separated by space
        new_name: e.g. if want to change name of time series
        description: e.g. if want to change name of time series
        overwrite_tag: e.g. set to true if you want to overwrite all the tags. Else will just append to current tags. valid tags must have 2 component. 1. period:day 2. freq:undefined. Both must be in to make sure updates works
        take_control: if true, it will take control of time series
        delete: if true will delete.
    '''
    if not logger:
        logger = logging.getLogger(__name__)
    #check for blank data
    if not name.strip():
        raise ValueError("Invalid time series name {}. It cannot be empty".format(name))


    #get information first
    url = "{}/dataedit/{}".format(automan_root,quote(name))
    r = requests.get(url, auth=HttpNegotiateAuth())
    html_data = r.content
    soup = BeautifulSoup(html_data, "html.parser")

    # logger.info("URL:{}".format(url))
    try:
        name_submit = soup.find("input",{"name":"name"})["value"]
        description_submit = soup.find("input",{"name":"description"})["value"]
        tags_old = soup.find("input",{"name":"oldtags"})["value"]
        tags_submit = tags_old
    except:
        err_msg="[automan_update_ts] error. The time series may not exists. TS Name: {} URL:{}".format(name,url)
        logger.error(err_msg)
        raise ValueError(err_msg)

    if tags and tags.strip(): #if tags is not none and tags is not empty
        if overwrite_tags:
            # print "==overwrite tags: {}".format(tags)
            tags_submit = tags
        else:
            # print "==dont overwrite tags: {}".format(tags)
            #split up the current tags
            tags_current_split = tags_submit.split(" ")
            #split up the existing tags
            tags_new_split = tags.split(" ")
            #add only tags that doesn't exists
            for tag_new_split in tags_new_split:
                if tag_new_split not in tags_current_split:#add only tags that doesn't exists
                    tags_submit = tags_submit+" "+tag_new_split


    if description and description.strip() and overwrite_description: #if valid description
        description_submit=description

    if new_name and new_name.strip(): #if valid description
        name_submit=new_name

    if "period:" not in tags_submit :
        raise ValueError("You must have one and only one'period!' tag: {}".format(tags_submit))


    payload = {"name":name_submit,"description":description_submit,"tags":tags_submit,"oldtags":tags_old,"save":"Submit Changes"}
    if take_control:
        payload['take']="Take Control"
        payload.pop("save",None)
    if delete:
        payload['delete']="Delete Series"
        payload.pop("save",None)
        payload.pop("take",None)
    # logger.info(url)
    # logger.info(payload)
    # print payload
    # print tags_submit
    # print tags_old
    r = requests.post(url, auth=HttpNegotiateAuth(),data=payload)

    return r

def automan_create_ts(name,description,tags="",period="sparse",overwrite_tags=False,overwrite_description=False,logger=None):
    '''
    Create time series in automan with the tags. If time series exists, it will just update the tags.
    Required:
        name: name of the series. Convention e.g. D_CHINAPRICES_SHANGHAI_REBAR. D=daily, ....
    Optional:
        description: The description of the time series
        tags: tags to add separated by space. e.g. "nymex year:2016"
        period: default is 'sparse'. can be day,sparse,half_week,week,month,year

    '''
    if not logger:
        logger = logging.getLogger(__name__)

    try:
        #check period
        valid_period = "day","sparse","half_week","week","month","year"
        if period not in valid_period:
            raise ValueError("Invalid period {}. Must be day,sparse,half_week,week,month,year".format(period))
        #check for blank data
        if not name.strip():
            raise ValueError("Invalid name {}. It cannot be empty".format(name))
        if not description.strip():
            raise ValueError("Invalid description {}. It cannot be empty".format(description))

        url = "{}/dataadd".format(automan_root)
        payload = {"name":name,"description":description,"period":period}
        r = requests.post(url, auth=HttpNegotiateAuth(),data=payload)
        # logger.info(payload)
        # logger.info(r.text)

        #add the tags...
        tags+=" period:{}".format(period)
        # print "TAGS:"+tags
        # print "overwrite_tags:"+str(overwrite_tags)
        # logger.info(
        #                 u"====UPDATING TS. name={},tags={},overwrite_tags={},description={},overwrite_description={}".format(
        #                     name,tags,overwrite_tags,description,overwrite_description
        #                 )
        #             )
        automan_update_ts(name,tags=tags,overwrite_tags=overwrite_tags,description=description,overwrite_description=overwrite_description,logger=logger)
    except:
        logger.error("[_automan_create_ts] error.")
        raise


def automan_get_grid_url(tags="nymex year:2016",start_date=date(2000,1,1)):
    '''
    Function return the URL given a series of tags.
    '''
    tags_list = tags.split(" ")
    tags_params = ""
    for tag in tags_list:
        tags_params = "{}t={}&".format(tags_params,tag)

    url = "{}/pivot/?{}{}&startDate={}".format(automan_root,tags_params,"",start_date.strftime("%Y-%m-%d"))
    return url

def automan_read_by_name(name, instance = "ot-metals", start_date=date(2000,1,1)):
    '''
    Function reads data from automan and return as dataframe.
    e.g. x = automan_read_by_name("PRICESTORE_MYSTEEL_STORE_Q235BSHANGHAITHIN")
    Required:
        name: name of single series to read.
    Optional:
        default date of 2000-01-01
    '''

    # automan_root = _automan_base.format(instance)

    url = "{}/pivot/?s={}&startDate={:%Y-%m-%d}&csv=true&long=true&mergeDateTime=true".format(
        automan_root,
        name,
        start_date)
    r = requests.get(url, auth=HttpNegotiateAuth()) #established session using ntlm auth

    return pd.read_csv(StringIO(r.text))

def read_excel_into_automan(
                            src_excel_path='//sgsing022m/sing8data/ETIA/DRY/FERROUS/automan/automan_central.xlsx',
                            start_date=date(2016,9,20),
                            end_date=date.today(),#NO USE NOW [YET TO IMPLEMENT]
                            header_last_row=1,#the row where header ends. e.g. first two rows are use for header, this value should be 2. If only first row use for header, thsi value is 1.
                            logger=None,
                            sheet_dict={
                                    "TSIIndex":[
                                                {"ts_name":"D_TSI_N_EU_HRC",'cell_col_id':'B','date_col_idx':0},
                                                {"ts_name":"D_TSI_S_EU_HRC",'cell_col_id':'C','date_col_idx':0},
                                                ],
                                }
                            ):
    '''
    Function reads data from excel file and write them into automan
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(src_excel_path,read_only=True)

    for key, value in sheet_dict.iteritems():
        logger.info("==Processing Sheet {}...".format(key))
        #Go through each sheet
        sheet_name = key
        ts_names = value # array of dictionary
        '''
        [
            {"ts_name":"D_TSI_N_EU_HRC",'cell_col_id':'B','date_col_idx':0}
            {"ts_name":"D_TSI_S_EU_HRC",'cell_col_id':'C','date_col_idx':0}
        ]
        '''

        ws=wb[sheet_name]
        for idx, val in enumerate(ts_names):
            ts_names[idx]["array"]=[] #create the empty array for each column

        date_array=[]

        #loop through each row to get the data
        for row in ws.rows:
            #loop through the rows..

            row_idx = row[0].row
            if row_idx <= header_last_row: #Ignore header. by default header_last_row is 1 (for spreadsheet that use only first row for header)
                pass
            else:#read the rows and store the value
                #a) Read date

                date_col_idx=ts_names[0]['date_col_idx']
                the_date = row[date_col_idx].value.date()

                if((the_date >= start_date) and (the_date < end_date)):
                    logger.info("====Sheet {}: Processing Date {}...".format(sheet_name,the_date))
                    date_array.append(the_date)
                #b) Read value only if bigger than start_date.
                    for idx, val in enumerate(ts_names):
                        # logger.info("------Processing TS: {}...".format(val['ts_name']))
                        ts_name=val
                        cell_col_excel_idx = ts_name['cell_col_id']
                        #add the value to the array
                        ts_name["array"].append(ws[cell_col_excel_idx+str(row_idx)].value)
                        # logger.info("------>Processed TS: {}!".format(val['ts_name']))

                    logger.info("====>Sheet {}: Processed Date {}!".format(sheet_name,the_date))
                elif(the_date >= end_date): #if too much exceed end_date we escape reading rows.
                    break

        #create the data frame
        data_dict={}
        data_dict['Date']=date_array
        for idx, val in enumerate(ts_names):
            ts_name=val
            data_dict[ts_name["ts_name"]]=ts_name["array"]
        df = pd.DataFrame(data_dict)

        logger.info("==Finish Processing Sheet {}!".format(sheet_name))


        #automan write the data frame
        logger.info("==Writing to Automan {}...".format(sheet_name))
        result = automan_write(df)
        logger.info("==Finish Writing to Automan {}!".format(sheet_name))

        # return df


def automan_exists_ts(name="AVAILABILITY DAYS OF COKING COAL AVERAGE INVENTORIES: DOMESTIC LARGE _AND_ MEDIUM-SIZED STEEL MILLS"):
    '''
    Return True if exists or False otherwise
    '''
    return_array = automan_search_series_by_name(name)

    if len(return_array) > 0 and return_array[0]==name:
        return True
    else:
        return False


def automan_search_series_by_name(
                                    ts_name='SOCIAL_INV_REBAR_SOUTHWEST'
                                  ):
    '''
    Function makes use of the autocomplete query to get a list of time series name in automan
    e.g. ts_name='SOCIAL_INV_REBAR_SOUTHWEST', return ['W_MYSTEEL_SOCIAL_INV_REBAR_SOUTHWEST']
    '''
    url='http://ot-metals.automan.cargill.com/autocomplete?term={}'.format(ts_name)
    headers = {'accept': 'application/json, text/javascript, */*; q=0.01','x-requested-with':'XMLHttpRequest'}
    r = requests.get(url, headers=headers, auth=HttpNegotiateAuth())
    return eval(r.content)

def automan_extract_data_dictionary(
                                    tags="metals -source:pricestore",
                                    output_path="C:\Users\j256377\Desktop\Metals\Automan Project\data dictionary\{}_data_dictionary".format(datetime.now().strftime('%Y_%m_%d')), #don't need extension
                                    mysteel_keyvalue_path="C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\mapping csv\\keyvalue_jax_2017_01_09.xlsx",
                                    logger=None
                                    ):
    '''
    Function will generate a data dictionary from the tags given.
    It will include:
        1. ts_name: Name of time series in Automan
        2. description: Description of the time series
        3. tags: the tags in string format separated by space. e.g. "tag1 tag2"
        4. source: the source of the ts.
        5. freq: the frequency of the time series. e.g. d,bw,

        Note: For source and freq to work, time series must follow the name convention of <FREQ>_<SOURCE>_<NAME>.
             e.g. D_MYSTEEL,
    '''

    if not logger:
        logger = logging.getLogger(__name__)


    #1. Get the URL from the tags
    url="http://ot-metals.automan.cargill.com/"
    get_str="?"
    tags=tags.split(" ")
    for tag in tags:
        get_str+="t={}&".format(tag)
    get_str=get_str[:-1]
    url=url+get_str



    r = requests.get(url, auth=HttpNegotiateAuth())
    r_data = StringIO(r.text)
    html_data = r_data
    soup = BeautifulSoup(html_data)
    rows = soup.find("table",{"id":"series"}).find_all("tr")
    rows=rows[1:] #ignore the header

    ts_name_list = []
    description_list = []
    description_list_utf8 = []
    tags_list = []
    source_list = []
    freq_list = []

    valid_freq_letters = ['d','w','bw','m','y']

    for row in rows:
        detail_td = row.find_all("td")[4]
        #1. Name
        ts_name = detail_td.find("div",{"class":"sname"}).find("a").text.replace(" ","_")
        ts_name_list.append(ts_name)

        #2. Description
        description = detail_td.find("p",{"class":"sdesc"}).text#.encode("utf-8")
        description_list.append(description)
        description_utf8 = detail_td.find("p",{"class":"sdesc"}).text.encode("utf-8")
        description_list_utf8.append(description)

        #3. Tags
        tags = detail_td.find("div",{"class":"stags"}).text.replace("\n"," ").strip()
        tags_list.append(tags)

        #4. Source and Freq
        freq,src= None,None
        ts_name_split = ts_name.split("_")
        freq,src = ts_name_split[0],ts_name_split[1]
        if freq.lower() not in valid_freq_letters: #if not valid freq, maybe didn't include, so we set src to first value
            src = ts_name_split[1]
            freq = None

        source_list.append(src)
        freq_list.append(freq)


    df = pd.DataFrame({
        "ts_name":ts_name_list,
        "description":description_list,
        "tags":tags_list,
        "source":source_list,
        "freq":freq_list,
    })


    df_utf8 = pd.DataFrame({
        "ts_name":ts_name_list,
        "description":description_list_utf8,
        "tags":tags_list,
        "source":source_list,
        "freq":freq_list,
    })

    #reorder
    ordered_cols=['ts_name','description','source','freq','tags']
    df = df[ordered_cols]
    df_utf8 = df_utf8[ordered_cols]

    #==Map to unique Mysteel ID if any...==
    if os.path.exists(mysteel_keyvalue_path):
        df_keyvalue = pd.read_excel(mysteel_keyvalue_path).ix[:,:4] #get only first 4 columns
        #join to get the mysteel id
        df_join_w_mysteel_key_value = df.merge(df_keyvalue,how='left',left_on="ts_name",right_on="TYPE")
        ordered_cols.insert(4,'ID')
        df=df_join_w_mysteel_key_value[ordered_cols].rename(columns={"ID":"mysteel_id"})

    else:
        logger.warning("MySteel KeyValue mapping file at {} is missing.".format(mysteel_keyvalue_path))




    if output_path:
        df_utf8.to_csv(output_path+".csv",index=False,encoding="utf-8")
        df.to_excel(output_path+".xlsx",index=False)

    return df

def automan_read_into_excel(
                            automan_name_array=['D_SINA_FXRATE_CLOSE_USD_CNY'],
                            columns_dict={
                                            "D_SINA_FXRATE_CLOSE_USD_CNY":
                                                {'ws_name':'FXRates','date_col_idx':0,'cell_col_id':'B'}
                                           },
                            start_date=date(2016,9,21),
                            end_date=date.today(),#NO USE NOW [YET TO IMPLEMENT]
                            target_excel_path='//sgsing022m/sing8data/ETIA/DRY/FERROUS/automan/automan_prices.xlsx',
                            logger=None, #logger to pass over if any
                            backup=True,#set to true if you want to store backup copy first...
                            ):
    '''
    Function reads data from automan and write into excel file.
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    #may need to increase size of recursion limit
    current_recursion_limit = sys.getrecursionlimit()
    sys.setrecursionlimit(current_recursion_limit*100)

    if backup:
        old_file_name=os.path.basename(target_excel_path)
        new_file_name=old_file_name.replace(".xlsx","-{}.xlsx".format(datetime.now().strftime("%Y-%m-%d_%H_%M_%S")))
        backup_path = os.path.dirname(target_excel_path)+"/backup/"+new_file_name
        _copy(target_excel_path,backup_path)

    #Read data from automan
    for ts in automan_name_array:
        #Open the Excel file
        wb = openpyxl.load_workbook(target_excel_path)

        logger.info("==Processing Time Series {}...==".format(ts))
        df=automan_read_by_name(name=ts,start_date=start_date)
        df['Date']=pd.to_datetime(df['Date'],format="%d-%b-%Y")
        #if error thrown 'ValueError: time data '25 Oct 2016' does not match format '%d-%b-%Y' (match)' likely the name of the time series is wrong hence return empty set

        df=df.dropna()

        ws_name=columns_dict[ts]['ws_name']
        ws=wb[ws_name]
        date_col_idx=columns_dict[ts]['date_col_idx']
        cell_col_id=columns_dict[ts]['cell_col_id']

        row_to_add_row_id=None
        for index, row in df.iterrows():
            the_date = row['Date'].date()


            result = _find_row_to_add_by_date(ws=ws,date_col_idx=date_col_idx,the_date=the_date,always_return_row_id=True,fill_formulae=False)#set previous_row_t_diff_day to 0, so it will overwrite existing values.
            row_to_add_row_id = result['row_to_add_row_id']

            # if index==0: #for the first row value we find the row to insert, subsequent row we just append to previous row
            #     result = _find_row_to_add_by_date(ws=ws,date_col_idx=date_col_idx,the_date=the_date,previous_row_t_diff_day=0,always_return_row_id=True,fill_formulae=False)#set previous_row_t_diff_day to 0, so it will overwrite existing values.
            #     row_to_add_row_id = result['row_to_add_row_id']
            # else:
            #     row_to_add_row_id = row_to_add_row_id+1

            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id=get_column_letter(date_col_idx+1),the_date=the_date)
            cell = '{}{}'.format(cell_col_id,str(row_to_add_row_id));
            ws[cell] = row[ts]
            ws[cell].number_format = ws['{}{}'.format(cell_col_id,str(row_to_add_row_id-1))].number_format
            ws[cell].alignment = ws['{}{}'.format(cell_col_id,str(row_to_add_row_id-1))].alignment
        logger.info("==>Processed Time Series {}!==".format(ts))

        #save the file for each time series
        logger.info("==Saving Time Series data to file {}...==".format(target_excel_path))
        wb.save(target_excel_path)
        logger.info("==>Saved Time Series data to file {}!==".format(target_excel_path))


    #set back recursion limit
    sys.setrecursionlimit(current_recursion_limit)



def automan_read(tags="nymex year:2016",start_date=date(2000,1,1),end_date=date(2100,1,1),wide_table=True,convert_to_date=True):
    '''
    Function reads data from automan and return as dataframe.
    Required:
        tags: list of tags separated by space.
    Optional:
        default date of 2000-01-01
        wide_table: default True. If set to false, it will melt and return tall table
        convert_to_date: default False. Set to true if you want to return 'Date' column as date time object instead of string
    '''
    logger = logging.getLogger(__name__)
    #established session using ntlm auth
    try:
        # session = requests.Session()
        # session.auth = HttpNtlmAuth(_session_data['ntlm']['username'],_session_data['ntlm']['password'], session)

        tags_list = tags.split(" ")
        tags_params = ""
        for tag in tags_list:
            tags_params = "{}t={}&".format(tags_params,tag)

        url = "{}/pivot/?{}{}&startDate={}&endDate={}".format(automan_root,tags_params,"csv=true&long=true&mergeDateTime=true",start_date.strftime("%Y-%m-%d"),end_date.strftime("%Y-%m-%d"))
        # print url
        # r = session.get(url)
        r = requests.get(url, auth=HttpNegotiateAuth())

        r_data = StringIO(r.text)
        df = pd.read_csv(r_data)

        if not wide_table:
            column_names = list(df.columns)
            column_names.remove('Date')
            df = pd.melt(df, id_vars=["Date"], value_vars=column_names,var_name="price")

        if convert_to_date:
            col_date = 'Date'
            df[col_date]=pd.to_datetime(df[col_date],format="%d-%b-%Y",errors='ignore')

        return df
    except Exception, e:
        logger.error("[automan_read] NTLM authentication may not be completed. Try 'ai.setup_ntlm()' first.")
        raise


def automan_write(df=pd.DataFrame({
                                    "Date": [date(2016,8,14),date(2016,8,15)],
                                    "jax_test": [300,400],
                                    "jax test 2": [500,600],
                                    })
                  ):
    '''
    Function write data to automan. df is the input in terms of dataframe. Return the post request object
    Required:
        df: the data frame with first column "date" and the rest is the series name
        e.g.
                  Date  jax test 2  jax_test
        0  2016-08-12         100       100
        1  2016-08-13         200       200
    Optional:
    '''
    logger = logging.getLogger(__name__)
    try:
        #established session using ntlm auth
        # session = requests.Session()
        # session.auth = HttpNtlmAuth(_session_data['ntlm']['username'],_session_data['ntlm']['password'], session)

        s = StringIO()
        df.to_csv(s,line_terminator="\r\n",index=False)
        # df.to_csv("c:\\tmp\\tmp.csv",line_terminator="\n",index=False)
        df_in_str = s.getvalue()



        url = "{}/importvalues".format(automan_root)
        payload = {"data":df_in_str,"import":True,"importByKey":True}
        # r = session.post(url,data=payload)
        r = requests.post(url, auth=HttpNegotiateAuth(),data=payload)

        return r
    except Exception, e:
        logger.error("[automan_read] NTLM authentication may not be completed. Try 'ai.setup_ntlm()' first.")
        raise


def setup_ntlm():
    '''
    Run this function to read the ntlm_auth.p generated by the script generate_ntlm_auth.py
    '''
    logger = logging.getLogger(__name__)
    try:
        global _session_data
        from os.path import abspath
        file_path =  abspath('{}\\ntlm_auth.p'.format(os.path.dirname(__file__)))
        ntlm_auth  = (pickle.load(open( file_path, "rb" )))

        ntlm_auth_dic = {
                      "username"  : ntlm_auth["username"],
                      "password" : _decrypt(ntlm_auth['password']),
                    }
        _session_data['ntlm'] = ntlm_auth_dic
    except:
        logger.warn("[_setup_ntlm] Error reading ntlm_auth.p ")
        raise



def load_single_time_series_automan(
                                    mode="excel",
                                    opts={"path":"//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx"},
                                    time_series_name="jax_test",
                                    start_date = None,
                                    end_date = None
                                   ):
    '''
    Function will load time series into automan from various sources as determined by mode
    e.g.
        load_single_time_series_automan(
            mode="excel",
            opts={"path":"//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx"},
            time_series_name="ChinaPrices_Tianjin_Rebar",
            opts = {
                    "path":"//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx",
                    "ws":"ChinaPrices","date_col_letter_idx":"C","time_series_letter_idx":"F",
                    "start_row_excel_idx":609,"end_row_excel_idx":None,
                    "path":"//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx",
                    "ws":"ChinaPrices"
                }
        )
    Precondition:
        Required:
            mode:
                "excel": In excel mode, the opts is required with the parameters below:
                        opts["path"]: path of the excel e.g. "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx"
                        opts["ws"]: name of the worksheet e.g. "ChinaPrices"
                        opts["date_col_letter_idx"]: the col index with the date. start with "A" e.g. "C"
                        opts["time_series_letter_idx"]: the col index with the data. start with "A" e.g. "F"
                        opts["start_row_excel_idx"]: row number to start. default is 2 (ignore 1 which is the header). idx start from 1
                        opts["end_row_excel_idx"]: row number to end. Include this row id. default is None meaning it will read all rows... idx start from 1
                "df": In dataframe mode, opts['df'] consists of the data frame to load to automan
            time_series_name: the name of the time series in automan

        Optional:
            start_date: the start date to ingest. default None. If None, it will not check if date is earlier than start_date. Else ignore date that is earlier than start_date
            end_date: the end  date to ingest. default None. If None, it will not check if date is later than end_date. Else ignore date that is later than end_date
    Postcondition:
        return True if successful or False if otherwise
    '''
    logger = logging.getLogger(__name__)
    if mode=="excel":

        #set default parameters
        default_opts = {
                            "path":"//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx",
                            "ws":"ChinaPrices","date_col_letter_idx":"A","time_series_letter_idx":"B",
                            "start_row_excel_idx":1,"end_row_excel_idx":None,
                        }
        tmp_opts = default_opts.copy()
        tmp_opts.update(opts)
        opts = tmp_opts

        logger.debug(u"[load_single_time_series_automan] Reading excel file...{}".format(opts['path']))
        wb = openpyxl.load_workbook(opts['path'],data_only=True)#,keep_vba=True)
        logger.debug(u"[load_single_time_series_automan] Loaded excel file! {}".format(opts['path']))
        ws = wb[opts['ws']]

        num_rows = len(ws.rows)
        last_row_num =  num_rows + 1

        if opts["end_row_excel_idx"] and opts["end_row_excel_idx"] <= num_rows:
            last_row_num = opts["end_row_excel_idx"]+1

        date_array = []
        time_series_array = []

        #go through each row and store the date and time series
        for row_excel_idx in xrange(opts["start_row_excel_idx"],last_row_num):
            curr_date = ws[opts['date_col_letter_idx']+str(row_excel_idx)].value

            ignore =False
            end_loop = False
            if start_date:
                if isinstance(curr_date,datetime) and curr_date.date() < start_date :
                    ignore = True
            if end_date:
                if isinstance(curr_date,datetime) and curr_date.date() >= end_date :
                    ignore = True
                    end_loop = True

            if not ignore:
                if ws[opts['time_series_letter_idx']+str(row_excel_idx)].value is None:
                    logger.warn ("[load_single_time_series_automan] Empty value at Row {} | Date:{} |  Value:{}".format(str(row_excel_idx),ws[opts['date_col_letter_idx']+str(row_excel_idx)].value,ws[opts['time_series_letter_idx']+str(row_excel_idx)].value))
                else:
                    date_array.append(ws[opts['date_col_letter_idx']+str(row_excel_idx)].value)
                    time_series_array.append(ws[opts['time_series_letter_idx']+str(row_excel_idx)].value)

            if end_loop:
                break #stop reading the row if not needed

        #package to dataframe
        df = pd.DataFrame({
            "Date":date_array,
            time_series_name: time_series_array,
        })

        #remove missing Value
        df = df[pd.notnull(df[time_series_name])]

        count_row = df.shape[0]
        logger.debug("[load_single_time_series_automan] Loading {} rows into Automan...".format(count_row))
        load_single_time_series_automan(mode="df",opts={"df":df},time_series_name=time_series_name)
        logger.debug("[load_single_time_series_automan] Loaded {} rows into Automan!".format(count_row))

        return df

    elif mode == "df":
        if ("df" not in opts):
            err_msg = "DataFrame is not found in opts['df']"
            logger.error(err_msg)
            raise ValueError(err_msg)
        else:
            df = opts['df']
            result = automan_write(df)

            return True

    else:
        return False


    pass

def setup_proxy():
    logger = logging.getLogger(__name__)
    try:
        global proxyDict
        from os.path import abspath
        file_path =  abspath('{}\\proxy_setting.p'.format(os.path.dirname(__file__)))
        print file_path
        http_proxy  = _decrypt(pickle.load(open( file_path, "rb" )))
        https_proxy = http_proxy
        proxyDict = {
                      "http"  : http_proxy,
                      "https" : https_proxy,
                    }
    except:
        logger.warn("[setup_proxy] Proxy settings not detected. Try generating the file using generate_proxy_link.py")
        raise

def _encrypt(msg,key="78nsGs3tosd83jfs"):
    obj = AES.new(key, AES.MODE_CFB, 'This is an IV456')
    return obj.encrypt(msg)

def _decrypt(encrypted_str,key="78nsGs3tosd83jfs"):
    obj = AES.new(key, AES.MODE_CFB, 'This is an IV456')
    return obj.decrypt(encrypted_str)

def parse_date_arg(argv,return_last_biz_day_default=True):
    '''
    Pass in the system argv and return a date object to be used by each task function. Date must be passed in "YYYY-MM-DD" format e.g. 2016-01-31
    If argv is empty, last biz day is returned when return_last_biz_day_default is True, else None is return
    '''
    if not argv:
        logger.info("[parse_date_arg] No date agrument present")
        return None
    logger = logging.getLogger(__name__)
    if len(argv) < 2:
        if return_last_biz_day_default:
            return get_biz_date(-1)
        else:
            return None
    else:
        try:
            the_date = datetime.strptime(str(argv[1]).strip(),"%Y-%m-%d").date()
            logger.info("[parse_date_arg] Date argument received. Processing for the date {}".format(the_date))
            return the_date
        except Exception, e:
            logger.warn("[parse_date_arg] Invalid date format {}. Setting date to today {}".format(argv[1],get_biz_date(0)))
            return get_biz_date(-1)

def _change_my_steel_page_html(url,pg_num):
    '''
    Fit in a URL like http://list1.mysteel.com/article/p-3571-------------1.html
    and return a new url with the new page number
    '''
    #return re.sub("[0-9]*\.html","2.html",'http://list1.mysteel.com/article/p-3571-------------1.html')
    return re.sub("[0-9]*\.html","{}.html".format(pg_num),url)

def _get_domain_from_url_str(url,trailing_slash=False):
    '''
    Function return domain from url e.g. url = "https://www.thesteelindex.com/en/restricted-area",
                                    domain = "https://www.thesteelindex.com/ [if trailing_slash is True]
                                    domain = "https://www.thesteelindex.com [if trailing_slash is False]
    '''

    parsed_uri = urlparse(url)
    if trailing_slash:
        domain = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)
    else:
        domain = '{uri.scheme}://{uri.netloc}'.format(uri=parsed_uri)
    return domain

def insert_rows(ws, row_idx, cnt, above=False, copy_style=True, fill_formulae=True):
    '''
    Summary: Function insert row at specified row index.
            https://bitbucket.org/snippets/openpyxl/qyzKn/sample-code-for-inserting-rows-in-an
            Usage:
            * insert_rows(2, 10, above=True, copy_style=False)

    Author: Jax
    Date created: 14 Jul 2016
    Date last updated: 14 Jul 2016
    Precondition:
        Required:
                :param row_idx: Row index specifying where to insert new rows.
                :param cnt: Number of rows to insert.
                :param above: Set True to insert rows above specified row index.
                :param copy_style: Set True if new rows should copy style of immediately above row.
                :param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.
        Optional:
            t: t is an integer mark by day. It will return today's date + t. t can be negative for previous day.
    Postcondition:
        Required: N/A
        Optional: N/A
    '''

    CELL_RE  = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    row_idx = row_idx - 1 if above else row_idx

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$",""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas   = set()
    new_cells = dict()
    new_fas   = dict()
    for c in ws._cells.values():

        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == Cell.TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in ws.formula_attributes and 'ref' in ws.formula_attributes[old_coor]:
                ws.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    ws.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row,c.col_idx))
            c.row += cnt
            new_cells[(c.row,c.col_idx)] = c
            if old_coor in ws.formula_attributes:
                old_fas.add(old_coor)
                fa = ws.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del ws._cells[coor]
    ws._cells.update(new_cells)

    for fa in old_fas:
        del ws.formula_attributes[fa]
    ws.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    for row in range(len(ws.row_dimensions)-1+cnt,row_idx+cnt,-1):
        new_rd = copy.copy(ws.row_dimensions[row-cnt])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd
        del ws.row_dimensions[row-cnt]

    # Now, create our new rows, with all the pretty cells
    row_idx += 1
    for row in range(row_idx,row_idx+cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(ws.row_dimensions[row-1])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd
        for col in range(1,ws.max_column):
            col = get_column_letter(col)
            cell = ws.cell('%s%d'%(col,row))
            cell.value = None
            source = ws.cell('%s%d'%(col,row-1))
            if copy_style:
                cell.number_format = source.number_format
                cell.font      = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border    = source.border.copy()
                cell.fill      = source.fill.copy()
            if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in ws.formula_attributes and 'ref' not in ws.formula_attributes[s_coor]:
                    fa = ws.formula_attributes[s_coor].copy()
                    ws.formula_attributes[cell.coordinate] = fa
                # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d"%(row - 1),
                    lambda m: m.group(1) + str(row),
                    source.value
                )
                cell.data_type = Cell.TYPE_FORMULA

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(ws.merged_cell_ranges):
        ws.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            cr
        )


    return ws





def get_date_str(t=0):
    '''
    Summary: Function will return a string of format YYYY-MM-DD e.g. 2016-01-01
    Author: Jax
    Date created: 14 Jul 2016
    Date last updated: 14 Jul 2016
    Precondition:
        Required: N/A
        Optional:
            t: t is an integer mark by day. It will return today's date + t. t can be negative for previous day.
    Postcondition:
        Required: N/A
        Optional: N/A
    '''
    the_day = date.today() + timedelta(t)
    the_day_str = the_day.strftime("%Y-%m-%d")
    return the_day_str



def get_date_str_frm_date(the_date,separator="-"):
    '''
    Summary: Function will return a string of format YYYY-MM-DD e.g. 2016-01-01 give a date/datetime object
    Author: Jax
    Date created: 18 Jul 2016
    Date last updated: 18 Jul 2016
    Precondition:
        Required:
            the_date: the date/datetime object
        Optional:
            separator: the separator e.g. "_" or "-". default is "-"
    Postcondition:
        Required: N/A
        Optional: N/A
    '''
    return the_date.strftime("%Y{0}%m{0}%d".format(separator))


def get_date(t=0,the_date=date.today()):
    '''
    Summary: Function will return the a date object
    Author: Jax
    Date created: 15 Jul 2016
    Date last updated: 15 Jul 2016
    Precondition:
        Required: N/A
        Optional:
            t: t is an integer mark by day. It will return today's date + t. t can be negative for previous day.
    Postcondition:
        Required: N/A
        Optional: N/A
    '''
    the_day = the_date + timedelta(t)
    return the_day




def is_bizday(t=date.today()):
    '''
    Summary: Function takes a take and return 'True' for working day, and 'False' for non-working day.
    Author: Jax
    Date created: 18 Jul 2016
    Date last updated: 18 Jul 2016
    Precondition:
        Required: t
    Postcondition:
        return a date object
    '''
    # print t
    if t.weekday() == 6 or t.weekday() == 5:
        return False
    else:
        return True


def get_biz_date(t=0,the_date=date.today()):
    '''
    Summary: Function will return the a business day (i.e. not sat/sun).
    If the t makes the day a non-business day,
    it will keep subtracting till a business day is reached when t is negative
    If t is positive, it will add till a business day is reached

    Author: Jax
    Date created: 15 Jul 2016
    Date last updated: 15 Jul 2016
    Precondition:
        Required: N/A
        Optional:
            t: t is an integer mark by day. It will return today's date + t. t can be negative for previous day.
            the_date: the reference date
    Postcondition:
        return the business day with respect to the_date.
    '''
    the_day = the_date + timedelta(t)
    if(is_bizday(the_day)): #if it is the business day we just return the day
        return the_day
    else: #else not working day...
        if(t>=0):
            return get_biz_date(t+1,the_date)
        else:
            return get_biz_date(t-1,the_date)



def _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=get_biz_date(-1)):
    '''
    Summary: Function will update a sheet date column row where date is the same as the_date
    Author: Jax
    Date created: 22 Jul 2016
    Date last updated: 22 Jul 2016
    Precondition:
        Required:
            row_to_add_row_id: the row to add the date
            ws: the worksheet of openpyxl
        Optional:
            date_col_id: the date column in string e.g. "A"
            the
    Postcondition:
        the ws will have the date updated
    '''
    date_cell = "{}{}".format(date_col_id,str(row_to_add_row_id))
    ws[date_cell]=the_date
    ws[date_cell].number_format = ws['{}{}'.format(date_col_id,str(row_to_add_row_id-1))].number_format


def run_macro_in_excel_file_and_save(
        excel_input_path="C:\Users\j256377\Desktop\Metals\Automan Project\\auto_wanda_mysteel\\automan_data.xlsm",
        macro_name="Module1.Update",
        visible=False, #to show excel file or not
        delay=10000 #in milliseconds. Estimate how long macro takes to run
    ):
    '''
    Function will
    1. Open the excel
    2. Run the macro of macro_name
    3. After that save the file
    '''
    if visible:
        visible_str = "True"
    else:
        visible_str = "False"
    run_vbs_script(
            script= """
                Dim xlApp, xlBook

                Set xlApp = CreateObject("Excel.Application")
                xlApp.DisplayAlerts = False
                xlApp.Application.Visible = {3}
                Set xlBook = xlApp.Workbooks.Open("{0}")



                xlApp.Run "{1}"


                WScript.Sleep {2}

                xlbook.Save
                xlBook.Close False
                set xlBook = Nothing

                xlApp.Quit
                Set xlApp = Nothing

                WScript.Echo "Finished Updating Automan."
                WScript.Quit
            """.format(excel_input_path,macro_name,delay,visible_str),
            tmp_path="C:\\tmp",delete_vb=True
        )

    pass

def run_vbs_script(
        script= """
        'Write Sheet's full path here
        strPath = "\\sgsing022m\sing8data\ETIA\DRY\Ferrous\Steel\China SnD\Property index\Wind 30 cities.xls"

         'Create an Excel instance and set visibility of the instance
        Set objApp = CreateObject("Excel.Application")
        objApp.Visible = True

         'Open workbook; Run Macro; Save Workbook with changes; Close; Quit Excel
        Set wbToRun = objApp.Workbooks.Open(strPath,True)
        wbToRun.Save
        wbToRun.Close
        objApp.Quit

        """,
        tmp_path="C:\\tmp",delete_vb=False
    ):
    '''
    Function run a vb script as per script.
    '''
    vbscript=script

    if not os.path.exists(tmp_path):
        os.makedirs(tmp_path)

    vbs_full_path = "{}/tmp_vbs.vbs".format(tmp_path)

    with open(vbs_full_path, "w") as vbscript_file:
        vbscript_file.write(vbscript)

    subprocess.call("cmd /c {}".format(vbs_full_path)) #this thread will stop untill subprocess is completed

    if delete_vb:
        os.remove(vbs_full_path)



def xls_conversion_vbscript(src,dest,tmp_path="C:\\tmp",delete_vb=False):
    '''
    Summary: Function eventually converts a xlsb,xlsx,xlsm to a xlsm file. It utilizes a VBScript as in between.
             It returns the path in string of the converted file (local) which is essentially 'dest'.
    Author: Jax
    Date created: 19 Jul 2016
    Date last updated: 19 Jul 2016
    Precondition:
        Required:
            src:
                Src of the xlsb/xlsx file:. e.g. \\sgsing022m\sing8data\Metals Risk Team\Metals_risk_data\PriceStore\PriceStore.xlsb
            dest:
                destination of the xlsm file: e.g. 'C:/Users/j256377/Desktop/Metals/AutoIntern/tmp_data/paper_price/PriceStore.xlsm'
    Postcondition:
        return path of the new xlsm file stored locally
    '''

    src=src.replace('/','\\')
    dest=dest.replace('/','\\')


    #create the directory for the destination and tmp
    directory = os.path.dirname(dest)
    if not os.path.exists(directory):
        os.makedirs(directory)

    if not os.path.exists(tmp_path):
        os.makedirs(tmp_path)


    vbscript = """
        dim filesys
        Set filesys = CreateObject("Scripting.FileSystemObject")

        set objExcel = createObject("Excel.Application")
        objExcel.visible = False
        objExcel.EnableEvents = False

        set objWb = objExcel.Workbooks.Open("{}")

        If filesys.FileExists("{}")  Then
          filesys.DeleteFile("{}")
        End If

        objWb.saveas "{}",52
        objWb.close

    """.format(src,dest,dest,dest)

    convertor_full_path = "{}/xls_convertor_tmp.vbs".format(tmp_path)

    with open(convertor_full_path, "w") as vbscript_file:
        vbscript_file.write(vbscript)

    subprocess.call("cmd /c {}".format(convertor_full_path)) #this thread will stop untill subprocess is completed

    if delete_vb:
        os.remove(convertor_full_path)

    return dest


def _find_row_to_add_by_date(ws,date_col_idx=0,the_date=get_biz_date(-1),previous_row_t_diff_day=-1,always_return_row_id=False,logger=None,fill_formulae=True):
    '''
    Summary: Function aim is to find the row number to insert the new date of the_date.
             It will return the row id.
             Go through each row looking at the date_col.
            1) Once the row with the date == to the previous biz day of the_date,
                a) A new row will be added below this row (if the next row of the_date is not found)
                b) the row id return will be its row_id + 1.
            2) Else If the exact date of the_date is found, it will return that row id. No new row will be inserted!
            3) Else if the next date is bigger than the_date, it will return the row before that and insert a new row

            If 1) or 2) is not fufilled, it will raise error when always_return_row_id is set to False. Else it will always return last row id + 1

    Author: Jax
    Date created: 19 Jul 2016
    Date last updated: 21 Jul 2016
    Precondition:
        Required:
        Optional:
            ws:
                worksheet
            date_col_idx:
                the col index with the date. start with 0
            the_date:
                the date to insert new data. E.g. if date to insert is 19 Jul 2016(Tue), it will look for 18 Jul 2016 (Mon).
            previous_row_t_diff_day:
                the date of the previous row difference in terms of day with the date. e.g. if previous_row_t_diff_day=-7, the_date is 22 Jul, the previous row to look for is 17 Jul
            always_return_row_id:
                default False. But if set to true, if this function cannot find a date, it will just return the last row id + 1 and not raise error.
            fill_formulae:
                Default is true. Set to true if you want to copy formulae from row.
    Postcondition:
        return {"ws", "row_to_add_row_id"}
            ws: the inserted row ws
            row_to_add_row_id: the row_id to add. e.g. row id begins with 1. It is the excel row id.
    '''
    row_to_add_row_id = None
    row_found = False
    row_the_date_found = False
    row_total_num = len(ws.rows)

    yesterday = get_biz_date(t=previous_row_t_diff_day,the_date=the_date)

    if not logger:
        logger = logging.getLogger(__name__)

    for row in ws.rows:
        last_row =False
        if row[0].row == row_total_num:
            last_row=True

        if not row_found:
            #if the previous day date is found and last row...
            if isinstance(row[date_col_idx].value,datetime) and row[date_col_idx].value.date() == yesterday and last_row:
                row_to_add_row_id = row[date_col_idx].row + 1
                break
            #elif the current date is found...No need insert new row
            elif isinstance(row[date_col_idx].value,datetime) and row[date_col_idx].value.date() == the_date:
                row_to_add_row_id = row[date_col_idx].row
                row_the_date_found=True #set true so that we don't create new row
                break
            #elif we found a date later than this date, this is the row to fill...
            elif isinstance(row[date_col_idx].value,datetime) and row[date_col_idx].value.date() > the_date:
                row_to_add_row_id = row[date_col_idx].row
                break #row is found so we break from looping rows

            # if row_to_add_row_id: #this is set when the yesterday of the_date is found ...
            #     if isinstance(row[date_col_idx].value,datetime) and row[date_col_idx].value.date() == the_date:
            #         row_the_date_found = True
            #         break

    if not row_to_add_row_id and not always_return_row_id: #cannot find the date
        raise ValueError("[_find_row_to_add_by_date] cannot find the row to insert date {}".format(get_date_str_frm_date(the_date)))
    elif not row_to_add_row_id and always_return_row_id:
        logger.warn("[_find_row_to_add_by_date] previous date or eact date not found. Returning the last row {}.".format(row[date_col_idx].row))
        row_to_add_row_id = row[date_col_idx].row+1
        row_the_date_found=True

    if not row_the_date_found:
        #insert rows below (which will copy previous row as well...)
        insert_rows(ws=ws,row_idx=row_to_add_row_id-1,cnt=1,fill_formulae=fill_formulae)


    return {"ws":ws,"row_to_add_row_id":row_to_add_row_id}

def _copy(src,dest):
    '''
    This copy function will create directory at the same time for the destination
    '''
    directory = os.path.dirname(dest)
    if not os.path.exists(directory):
        os.makedirs(directory)
    shutil.copy(src ,dest)


def _equal_chinese_char(str1,str2):
    '''
    str1 is the original UTF text typed in the python file
    str2 can be any encoding typically from beautiful soup.
    '''
    str1 = str1.strip()
    str2 = str2.strip()

    equal1 = False
    equal2 = False
    equal3 = False
    equal4 = False
    equal5 = False
    try:

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            equal1 = (str1 == str2.encode('utf-8'))
    except Exception as e:
        pass

    try:
        equal2 = str1.decode("utf-8") == str2
    except Exception as e:
        pass

    try:
        equal3 = unicode(str1.decode("utf-8").encode("gb2312"), encoding='unicode_escape') == str2
    except Exception as e:
        pass

    try:
        equal4 = unicode(str1.encode("gb2312"), encoding='unicode_escape') == str2
    except Exception as e:
        pass

    try:
        equal5 = str1.encode("utf-8") == str2.encode('utf-8')
    except Exception as e:
        pass


    success = (
        equal1
        or equal2
        or equal3
        or equal4
        or equal5
    )
    #
    # if not success:
    #

    return success



def _in_chinese_char(str1,str2):
    '''
    str1 is the original UTF text typed in the python file
    str2 can be any encoding typically from beautiful soup.
    '''
    str1 = str1.strip()
    str2 = str2.strip()

    equal1 = False
    equal2 = False
    equal3 = False
    equal4 = False
    equal5 = False
    try:
        equal1 = (str1 in str2.encode('utf-8'))
    except Exception as e:
        pass

    try:
        equal2 = str1.decode("utf-8") in str2
    except Exception as e:
        pass

    try:
        equal3 = unicode(str1.decode("utf-8").encode("gb2312"), encoding='unicode_escape') in str2
    except Exception as e:
        pass

    try:
        equal4 = unicode(str1.encode("gb2312"), encoding='unicode_escape') in str2
    except Exception as e:
        pass

    try:
        equal5 = str1.encode("utf-8") in str2.encode('utf-8')
    except Exception as e:
        pass

    success = (
        equal1
        or equal2
        or equal3
        or equal4
        or equal5
    )
    #
    # if not success:
    #

    return success



def _get_mysteel_price(
            url="http://list1.mysteel.com/article/p-3571------0-0-0-----1.html", #page 2 is just http://list1.mysteel.com/article/p-3571------0-0-0-----1.html
            the_date=get_biz_date(-1),
            session_requests=None,
            mode = "furnace_raw",
            price_name = None,
            spec="Ф16-25",
            origin="东华钢铁",
            product="螺纹钢",
            area="上海",
            suppress_exception = False #set to True if you don't want this function to raise exeception
    ):
    '''
    Get price from mysteel
    Author: Jax
    Date created: 19 Jul 2016
    Date last updated: 21 Jul 2016
    Precondition:
        Required:
        Optional:
            url:
                url of the mysteel first page. default http://list1.mysteel.com/article/p-3571------0-0-0-----1.html
            the_date:
                date of getting the mysteel price. Default the_date=get_biz_date(-1),
            mode:
                mode can be either:
                "furnace_raw": get the 出厂 price blast furnance raw material 炉料 http://list1.mysteel.com/article/p-3571-------------1.html
                "construction": construction material 建材 e.g url = "http://list1.mysteel.com/market/p-228-----010101-0-0104-------1.html"
                    in this mode, additional parameters needed as below
                        "price_name" in chinese is needed for navigating to the page with the price. e.g. "天津市场建筑钢材价格行情" is needed to search for the title
                        "product" in chinese e.g. "螺纹钢"
                        "spec" 规格（mm）e.g.
                        "origin" 钢厂/产地" e.g. "东华钢铁"
                "hrc"/"crc": HRC 热轧 first page e.g. "http://list1.mysteel.com/price/p-10058--010103--1.html"
                "crc": CRC 冷轧 first page e.g. "http://list1.mysteel.com/price/p-10053--010104--1.html"
                    area: e.g. "上海" col
                    spec: e.g. "1.0mm" row
    Postcondition:
        return the price (float). return None if data not found or not ready. Will not raise error if missing data. Just return none.
    '''
    logger = logging.getLogger(__name__)

    if(not session_requests):
        session_requests = _get_login_session(type="mysteel")

    try:

        the_link=None
        max_pg_to_find = 10
        current_pg = 1
        while(not the_link and current_pg < max_pg_to_find):

            url = _change_my_steel_page_html(url,current_pg)
            #fetch the index page
            result = _retry_request(url=url,encoding="gb2312",session_requests=session_requests)
            html_data = result.content
            soup = BeautifulSoup(html_data,from_encoding="gb2312")
            links = soup.find("ul",{"class":"nlist"}).find_all("li",class_=lambda x: x != 'dashed')

            for link in links:
                link_date = datetime.strptime(link.find("span",{"class":"date"}).text,"%Y-%m-%d %H:%M").date()

                if(
                    (not price_name)#if price_name is None we don't have to match..
                    or
                    _in_chinese_char(price_name,link.text)
                    # price_name.encode('utf-8') in link.text.encode('utf-8')
                    # or
                    # price_name.decode("utf-8") in link.text
                    # or
                    # unicode(price_name.decode("utf-8").encode("gb2312"), encoding='unicode_escape') in link.text.strip()
                ):
                    if link_date == the_date:
                        logger.info("[_get_mysteel_price] Found the link in page {} for {}. ".format(current_pg,url))
                        the_link = link.find("a").get("href")
                        break #found the link!
                    elif link_date < the_date: #if the page doesn't have the latest figures, stop searching
                        logger.warn("[_get_mysteel_price] missing price {} for this date {}. The max date in the index page is {}".format(url,the_date,link_date))
                        return None #data not found


            current_pg = current_pg+1
            if not the_link:
                logger.info("[_get_mysteel_price] Couldn't find in page {} for {}. Trying the page {}".format(current_pg-1,url,current_pg))

        if not the_link:
            logger.warn("[_get_mysteel_price] couldn't find link in {} for the date {}".format(url,the_date))
            return None
        #a) Furnace raw
        if mode == "furnace_raw":

            #go to the link and find 出厂 price
            result = _retry_request(url=the_link,encoding="gb2312",session_requests=session_requests)
            html_data = result.content
            soup = BeautifulSoup(html_data)

            try:
                price = float(soup.find("div",{"id" : "text"}).find("table").find_all("tr")[2].find_all("td")[-1].find("span").contents[0])
                logger.debug("[_get_mysteel_price] price is {}".format(str(price)))
                return price
            except:
                logger.warn("[_get_mysteel_price] missing price {} for this date {}".format(url,the_date))
                return None


        #b) construction material
        elif mode == "construction":

            #go to the link and find the price
            result = _retry_request(url=the_link,encoding="gb2312",session_requests=session_requests)
            html_data = result.content
            html_data = html_data.decode("gb2312","ignore")
            soup = BeautifulSoup(html_data,from_encoding="gb2312")

            rows = soup.find("table",{"id":"marketTable"}).find_all("tr")

            price = None
            num_row_to_skip = 2
            for idx,row in enumerate(rows):
                if idx >= num_row_to_skip:
                    row_product = row.find_all("td")[0].text.strip()
                    row_spec = row.find_all("td")[1].text.strip()
                    row_origin = row.find_all("td")[3].text.strip()
                    if(
                        # row_product.encode('utf-8') == product.encode("utf-8")
                        # and row_spec == spec.decode("utf-8")
                        # and row_origin == origin.decode("utf-8")

                        _equal_chinese_char(product,row_product)
                        and _equal_chinese_char(spec,row_spec)
                        and _equal_chinese_char(origin,row_origin)
                    ):
                        price = float(row.find_all("td")[4].text.strip())
                        break
            if not price:
                dump_path = "_get_mysteel_price_{}".format(datetime.now().strftime("%Y-%m-%d_%H_%M_%S"))
                logger.warn("[_get_mysteel_price] missing price {} for this date {}. Is the login ID still working? Check the dump at {}".format(url,the_date,dump_path))
                with open(dump_path, "w") as text_file:
                    text_file.write(u"{}.txt".format(html_data))
                # with codecs.open("{}_compared.txt".format(dump_path), "w", encoding="utf-8") as text_file:
                #     text_file.write("product:" + product.decode('utf-8') + ";spec:" +spec.decode('utf-8')+" ;origin:"+origin.decode('utf-8'))
                # with codecs.open(dump_path, "w", "gb2312") as text_file:
                #     text_file.write(html_data.encode('gb2312'))


            return price


        elif mode == "crc" or mode == "hrc":
            #go to the link
            result = _retry_request(url=the_link,encoding="gb2312",session_requests=session_requests)
            html_data = result.content
            soup = BeautifulSoup(html_data)

            rows = soup.find("table",{"id":"priceTable"}).find_all("tr")

            price = None
            col_idx = None
            row_idx = None
            for idx,row in enumerate(rows):
                if idx == 0: #for first row...
                    for cidx,td in enumerate(row.find_all("td")):
                        # if td.text.strip() == area.decode("utf-8"):
                        if _equal_chinese_char(area,td.text):
                            col_idx = cidx
                            break
                else:
                    if(
                        # row.find_all("td")[0].text.strip() == spec.decode("utf-8")
                        _equal_chinese_char(spec,row.find_all("td")[0].text)
                    ):
                        row_idx = idx
                        break

            if col_idx and row_idx:
                price = float(rows[row_idx].find_all("td")[col_idx].text.strip())
            else:
                logger.warn("[_get_mysteel_price] missing price {} for this date {}. Is the login ID still working?".format(url,the_date))
            return price


        else:
            raise ValueError("Invalid mode for _get_mysteel_price! Your mode is {}".format(mode))
    except Exception, e:
        msg = "[_get_mysteel_price] Error getting data from my steel. Is the login still working or the page has changed? Error msg: {}".format(str(e))
        if suppress_exception:
            logger.warning(msg)
        else:
            logger.error(msg)
            raise




def _get_sina_finance_fx(
            url="http://vip.stock.finance.sina.com.cn/forex/api/jsonp.php/var%20_fx_s{0}{1}=/NewForexService.getDayKLine?symbol=fx_s{0}&_={1}".format("[[CUR]]","[[DATE]]"),
            cur_from="usd",
            cur_to="cny",
            price_type="close", #can be 'open','lowest','close','highest'
            the_date=get_biz_date(-1),
            end_date=date.today(),
            suppress_exception=False,
    ):
    '''
    Get the exchange rate for two currency at a particular date.
    Author: Jax
    Date created: 19 Jul 2016
    Date last updated: 21 Jul 2016
    Precondition:
        Required:
        Optional:
            url:
                url of the finance.sina.com.cn e.g."http://vip.stock.finance.sina.com.cn/forex/api/jsonp.php/var%20_fx_s{0}{1}=/NewForexService.getDayKLine?symbol=fx_s{0}&_={1}".format("[[CUR]]","[[DATE]]")
            cur_from:
                the reference currency. Default cur_from="usd"
            cur_to:
                the converted currency. Default: cur_to="cny"
            price_type:
                Price , #can be 'open','lowest','close','highest'
            the_date=get_biz_date(-1),
            end_date=date.today()
    Postcondition:
        return the price in float
    '''
    logger = logging.getLogger(__name__)

    # url="http://vip.stock.finance.sina.com.cn/forex/api/jsonp.php/var%20_fx_susdcny2016_7_21=/NewForexService.getDayKLine?symbol=fx_susdcny&_=2016_7_21"

    cur_combined = "{}{}".format(cur_from,cur_to) #e.g. usdcny
    date_str = get_date_str_frm_date(end_date,"_")
    url = url.replace("[[CUR]]",cur_combined).replace("[[DATE]]",date_str)

    # url = "http://localhost/usdcny.txt"

    logger.debug("[_get_sina_finance_fx] Fetching exchange rate from {}...".format(url))
    # print url
    # r  = requests.get(url,headers=headers)
    r = _retry_request(url)
    data = r.text
    all_price_str = data[data.find("new String(")+12:data.find("\")")]
    # print all_price_str
    logger.debug("[_get_sina_finance_fx] Fetched exchange rate from {}!".format(url))

    prices = all_price_str.split('|')
    prices_array = []
    for price in prices:
        price = price.split(',')
        price_dict = {
                        "date": price[0],
                        "open": float(price[1]),
                        "lowest": float(price[2]),
                        "highest": float(price[3]),
                        "close": float(price[4])
                     }

        prices_array.append(price_dict)


    err_msg = "[_get_sina_finance_fx] Can't find FX for {0} date {1} and price type of {2}".format(cur_combined,get_date_str_frm_date(the_date),price_type)
    # print prices_array
    for price in reversed(prices_array):
        if price['date']==get_date_str_frm_date(the_date):
            if price_type in price:
                return price[price_type]
            else:
                if not suppress_exception:
                    logger.error(err_msg)
                    raise ValueError(err_msg)#by right should be able to find in the loop before
                else:
                    logger.warn(err_msg)
                    return

    err_msg = "[_get_sina_finance_fx] Missing FX for date {0}".format(get_date_str_frm_date(the_date))
    if not suppress_exception:
        logger.error(err_msg)
        raise ValueError(err_msg)#by right should be able to find in the loop before
    else:
        logger.warn(err_msg)


def _get_sina_finance_future_price(
        url='http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=RB&hy=RB0&breed=RB0&type=inner',
        the_date=date.today()-timedelta(1),
        suppress_exception=False
    ):
    '''
    Summary: Function pull closing price from finance.sina.com.cn
    Author: Jax
    Date created: 19 Jul 2016
    Date last updated: 19 Jul 2016
    Precondition:
        Required:
        Optional:
            the_date:
                the date of the data to get the numbers
            url:
                url of the finance.sina.com.cn e.g.http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=RB&hy=RB0&breed=RB0&type=inner
    Postcondition:
        return the price in float
    '''
    logger = logging.getLogger(__name__)

    date_str = get_date_str_frm_date(the_date)
    url_with_date = "{}&start={}&end={}".format(url,date_str,date_str)

    # logger.debug("[_get_sina_finance_future_price] Fetching future price from {}...".format(url_with_date))
    # r  = requests.get(url_with_date,headers=headers)
    # logger.debug("[_get_sina_finance_future_price] Fetched future price from {}!".format(url_with_date))

    r = _retry_request(url=url_with_date)

    data = r.text
    soup = BeautifulSoup(data)
    try:
        tr_history = soup.find("tr", {"class" : "tr_2"})
        return float(tr_history.find_all('div')[1].contents[0])
    except Exception, e:
        err_msg = "[_get_sina_finance_future_price] Error scraping from sina finance. Error msg: {}".format(str(e))
        if not suppress_exception:
            logger.error(err_msg)
            raise
        else:
            logger.warn(err_msg)




def _get_login_session(type="steelindex",debug=False):
    '''
    Summary: Function logge into website and return the session_requests
    Author: Jax
    Date created: 27 Jul 2016
    Date last updated: 27 Jul 2016
    Precondition:
        Required:
            "type": The type of logged in session to return. default is steelindex
            It can be one of the below:
                1) steelindex
                2) steelbb
        Optional:
    Postcondition:
        return session_requests object with login
    '''
    logger = logging.getLogger(__name__)

    delay = 5
    num_tries = 5
    payload = _session_data[type]
    session_requests = requests.session()

    if debug:
        ("[_get_login_session] Login to {}...".format(_session_data[type]["login_url"]))

    logger.debug("[_get_login_session] Login to {}...".format(_session_data[type]["login_url"]))
    success = False
    i = 0
    while not success or i < num_tries:
        try:

            result = session_requests.post(
            	_session_data[type]["login_url"],
            	data = payload,
                allow_redirects=False,
                headers=headers,
                proxies=proxyDict
            )
            success=True
            print result
            break
        except Exception, e:
            i = i + 1
            # print "_retry_request failed! Retrying for {}th time".format(str(i))
            logger.debug("_get_login_session failed! Retrying for {}th time. Error msg {}".format(str(i),str(e)))
            if debug:
                print("_get_login_session failed! Retrying for {}th time".format(str(i)))
            time.sleep(delay)
            continue

    logger.debug("[_get_login_session] Logged in to {}!".format(_session_data[type]["login_url"]))
    if debug:
        ("[_get_login_session] Logged in to {}!".format(_session_data[type]["login_url"]))
    return session_requests






def _get_steelbb_price(
        the_date=get_biz_date(-1),
        session_requests=None,
        url='https://www.steelbb.com/?PageID=93&series_id=23&period_type=5',
        suppress_exception = False
    ):
    '''
    Summary: Function pull price from the steelbb.com.
    Author: Jax
    Date created: 21 Jul 2016
    Date last updated: 21 Jul 2016
    Precondition:
        Required:

        Optional:
            url:
                url of the steelbb.cn e.g.https://www.steelbb.com/?PageID=93&series_id=23&period_type=5
    Postcondition:
        return the {"min":min_price,"max":max_price,"avg":avg price} in float.
        if date is not found on the website return None
    '''
    logger = logging.getLogger(__name__)

    # logger.debug("fetching {}".format(url))
    if(not session_requests):
        session_requests = _get_login_session(type="steelbb")
    # logger.debug("fetching {}".format(url))
    # result = session_requests.get(
    #     url,
    #     headers=headers
    # )
    result = _retry_request(url,session_requests)
    # logger.debug("fetched {}".format(url))

    html_data = result.content
    soup = BeautifulSoup(html_data)


    try:
        date_rows = soup.find("tr", {"id" : "spt_inner_row_1"}).find("td").find_all("div")

        col_idx = None
        for idx,val in enumerate(date_rows):
            has_b = val.find("b")
            if has_b:
                row_date_str = str(val.find("b").contents[0])
                row_date = datetime.strptime(row_date_str,"%d %b %y").date()
                if row_date == the_date:
                    # print row_date_str
                    col_idx = idx
                    break


        #if we can find the date...
        if col_idx:
            # logger.info()
            price_rows = soup.find("tr", {"id" : "spt_inner_row_2"}).find("td").find_all("div")
            price_range = price_rows[col_idx].contents[0].strip()
            price_ranges = price_range.split("-")

            price_low = float(price_ranges[0].strip())
            price_high = float(price_ranges[1].strip())

            return {"min":price_low,"max":price_high,"avg":(price_low+price_high)/2}
        else:
            warn_msg = "[_get_steelbb_price] cannot find the price for the date {}. Data could be not updated yet on steelbb. Try again tomorrow.".format(get_date_str_frm_date(the_date))
            logger.warn(warn_msg)
            return {"min":None,"max":None,"avg":None}
            # raise ValueError(warn_msg)


    except Exception, e:
        err_msg = "[_get_steelbb_price] scrap error! Has the subscription for steelbb expired? Error msg: {}".format(str(e))
        if not suppress_exception:
            logger.error(err_msg)
            raise
        else:
            logger.warn(err_msg)
            return {"min":None,"max":None,"avg":None}








def _retry_request(
        url,
        session_requests = None,
        mode = "GET",
        num_tries = 10,
        delay = 3,
        encoding = 'utf-8',
        debug = False,
        timeout = 300
    ):
    '''
    Summary: Function will keep retrying for num_tries time sleeping for 5 seconds in between each try
    Author: Jax
    Date created: 28 Jul 2016
    Date last updated: 28 Jul 2016
    Precondition:
        Required:
            url:
                url that you want to get/post
        Optional:
            mode:
                mode of the request can be "GET" or "POST"
            delay:
                the number of seconds to delay before each retry
            encoding:
                The encoding type of the doc. Default is UTF8. For mysteel it is "gb2312"
    Postcondition:
        return results of the request if didn't fail...
    '''
    logger = logging.getLogger(__name__)

    time.sleep(0.5)

    success = False
    i = 0
    while not success or i < num_tries:
        try:
            the_request = None
            if not session_requests:
            #a)Non session mode
                the_request = requests
            else:
            #b) Session mode
                the_request = session_requests



            result = None
            if mode == "POST":
                if debug:
                    print("[_retry_request] fetching [POST] {}".format(url))
                logger.debug("[_retry_request] fetching [POST] {}".format(url))
                result = the_request.post(url,headers=headers,timeout=timeout,proxies=proxyDict)
                result.encoding = encoding
                logger.debug("[_retry_request] fetched [POST] {}".format(url))
                if debug:
                    print("[_retry_request] fetched [POST] {}".format(url))
            else:
                if debug:
                    print("[_retry_request] fetching [GET] {}".format(url))
                logger.debug("[_retry_request] fetching [GET] {}".format(url))
                result = the_request.get(url,headers=headers,timeout=timeout,proxies=proxyDict)
                result.encoding = encoding
                logger.debug("[_retry_request] fetched [GET] {}".format(url))
                if debug:
                    print("[_retry_request] fetched [GET] {}".format(url))

            #Check to make sure the content is not empty
            if result.content.strip() == "":
                raise ValueError("[_retry_request] Empty results")
            elif "Bad Gateway" in result.content.strip():
                dump_path = "_retry_request_{}.txt".format(datetime.now().strftime("%Y-%m-%d_%H_%M_%S"))
                logger.warn("[_retry_request] Cargill Proxy error DUMP: {}".format())
                with open(dump_path, "w") as text_file:
                    text_file.write("{}".format(result.content))
            else:
                return result

        except Exception, e:
            i = i + 1
            # print "_retry_request failed! Retrying for {}th time".format(str(i))
            logger.debug("[_retry_request] _retry_request failed! Retrying for {}th time. Error msg: {}".format(str(i),str(e)))
            if debug:
                print("[_retry_request] _retry_request failed! Retrying for {}th time".format(str(i)))
            time.sleep(delay)
            continue
        break

    #Retry still fails!
    raise ValueError("[_retry_request] _retry_request still failed after retrying for {} times".format(str(num_tries)))





def _get_thesteelindex_price(
        the_date=get_biz_date(-1),
        session_requests=None,
        url='https://www.thesteelindex.com/en/?cid=46&sid=1',
        suppress_exception = False
    ):
    '''
    Summary: Function pull price from the thesteelindex.com.
    Author: Jax
    Date created: 21 Jul 2016
    Date last updated: 21 Jul 2016
    Precondition:
        Required:
        Optional:
            url:
                url of the thesteelindex.cn e.g.https://www.thesteelindex.com/en/?cid=46&sid=1
    Postcondition:
        return the price in float
    '''
    logger = logging.getLogger(__name__)

    if(not session_requests):
        session_requests = _get_login_session(type="steelindex")


    # result = session_requests.get(
    #     url
    # )
    result = _retry_request(url=url,session_requests=session_requests)

    html_data = result.content
    soup = BeautifulSoup(html_data)

    try:
        date_rows = soup.find("tr", {"id" : "DataRow1"}).find_all("td")
    except Exception, e:
        err_msg = "[_get_thesteelindex_price] error parsing TSI {}- Dump of content: {}. Error msg: {}".format(url,html_data,str(e))
        if not suppress_exception:
            logger.error(err_msg)
            raise
        else:
            logger.warn(err_msg)
            return None

    try:

        str_date = the_date.strftime("%d/%m/%Y")
        col_idx = None
        for idx,val in enumerate(date_rows):
            if str(val.contents[0]) == str_date:
                col_idx = idx
                break


        price_rows = soup.find("tr", {"id" : "DataRow3"}).find_all("td")

        #if we can find the date...
        if col_idx:
            # logger.info()
            return float(price_rows[col_idx].contents[0])
        else:
            err_msg = "[_get_thesteelindex_price] cannot find the price for the date {}".format(get_date_str_frm_date(the_date))
            logger.error(err_msg)
            raise ValueError(err_msg)
    except Exception, e:
        err_msg = "[_get_thesteelindex_price] Error getting price from the steel index. Is the login still working or has the page changed? Error msg: {}".format(str(e))
        if not suppress_exception:
            logger.error(err_msg)
            raise
        else:
            logger.warn(err_msg)
            return None



def _get_feigang_price(
        url='http://www.feigang.net/price.aspx?colid=120&area=&keyword=%E5%BC%A0%E5%AE%B6%E6%B8%AF',
        the_date=get_biz_date(-1),
        suppress_exception = False
    ):
    '''
    Summary: Function pull the_date today price from feigang.net
    Author: Jax
    Date created: 22 Jul 2016
    Date last updated: 22 Jul 2016
    Precondition:
        Required:
        Optional:
            the_date:
                the date of the data to get the numbers
            url:
                url of the feigang without the date input http://www.feigang.net/ e.g.http://www.feigang.net/price.aspx?colid=120&area=&keyword=%E5%BC%A0%E5%AE%B6%E6%B8%AF
    Postcondition:
        return the price in float
    '''
    logger = logging.getLogger(__name__)

    try:
        date_str = get_date_str_frm_date(the_date)
        date_str_input1 = get_date_str_frm_date(the_date-timedelta(5))
        url_with_date = "{}&dateinput1={}&dateinput2={}&searchtype=0".format(url,date_str_input1,date_str)

        logger.debug("[_get_feigang_price] Fetching price from {}...".format(url_with_date))
        # r  = requests.get(url_with_date,headers=headers)
        r = _retry_request(url_with_date)

        data = r.text
        soup = BeautifulSoup(data)
        link_list = soup.find_all("li", {"class" : "r-result"})
        link_date_str = link_list[0].find("h2").find("span").contents[0]
        link_date = datetime.strptime(link_date_str,"%Y-%m-%d %H:%M").date()

        #check if the date is correct...
        if(link_date == the_date):
            the_link = _get_domain_from_url_str(url_with_date)+link_list[0].find('a').get('href')
            # html = urllib2.urlopen(the_link)
            r = _retry_request(the_link)
            data = r.text
            logger.debug("[_get_feigang_price] Fetched price from {} -> {}!".format(url_with_date,the_link))
            # data = html.read().decode('utf-8', 'ignore')

            soup = BeautifulSoup(data)
            tr_rows = soup.find("table", {"class" : "ContentTb"}).find_all("tr")
            today_price_idx = None
            for idx,row in enumerate(tr_rows[0].find_all("th")):
                 if row.text.strip() == "今天价格".decode("utf-8"):
                     today_price_idx = idx
                     break
            if today_price_idx:
                return float(tr_rows[1].find_all("td")[today_price_idx].text.strip())
            else:
                err_msg = "[_get_feigang_price] Error. Cannot find today price in feigang! The table may changed. Check the link!"
                if not suppress_exception:
                    logger.error(err_msg)
                    raise ValueError(err_msg)
                else:
                    logger.warn(err_msg)
                    return None
        else:
                err_msg = "[_get_feigang_price] Error. Cannot find price for {} in feigang! Feigang may not have the data yet.".format(the_date)
                if not suppress_exception:
                    logger.error(err_msg)
                    raise ValueError(err_msg)
                else:
                    logger.warn(err_msg)
                    return None

    except Exception, e:
        err_msg = "[_get_feigang_price] Error getting price from the feigang. Hasas the page changed? Error msg: {}".format(str(e))
        if not suppress_exception:
            logger.error(err_msg)
            raise
        else:
            logger.warn(err_msg)
            return None





def _daterange(start_date, end_date):
    for n in range(int ((end_date - start_date).days)):
        yield start_date + timedelta(n)



def _get_price_store_future(
                            price_store_path="C:\\tmp_data\\rv\\PriceStore.xlsm",
                            start_date=date.today()-relativedelta(days=2), #include start date
                            end_date=date.today()-relativedelta(days=1), #exlcude end date
                            sheet_name = 'SHFE.RBCNY.Store',
                            asset_name = None,
                            t=2,
                            excel_col=None, #e.g. "B"
                            get_mapping=False, #set to true if you just want to get ts_name
                            wb=None, #workbook. if set, it will not open the wb
                            ):
    '''
    Get the Prices from price store of future t. e.g. if t = 2 and today is 1 Sep, we take Nov price
    Dataframe is returned with first column date and second column the price
    if t = -1, we get the active1 and if t = -2, we get active2
    if t = None, we get spot price based on excel_col.
    Required:
        price_store_path: price store path
        start_date: the date to start (including)
        end_date: the date to end (excluding)
        sheet_name: name of the sheet
        t : t months away from the date. Default 2. set to negative for active contract and None for spot price
        excel_col: the column to read when t=0. Use excel alphabets column e.g. "B".

    '''
    logger = logging.getLogger(__name__)
    if not wb:
        wb = openpyxl.load_workbook(price_store_path, read_only=True, data_only=True)

    if t!=None and t < -2:
        logger.error("Error in _get_price_store_future")
        raise ValueError("t cannot be less than -2!")
    try:
        if not asset_name:
            asset_name = sheet_name
        #find the correct column for each month and year
        #while(the_date)
        col_idx_by_mth = {}
        for single_date in _daterange(start_date, end_date):
            month_string = single_date.strftime("%Y%m")
            col_idx_by_mth[month_string]=None

        ws = wb[sheet_name]
        row_found=False
        date_and_future_prices= []#  consists of the list of dates and price




        if t!=None and t < 0: #For active contnuous futures...
            cell_col_id = None
            if get_mapping==False:#only if not get_mapping mode...
                for row in ws.rows:
                    row_idx = row[0].row
                    if row[0].row == 1: #for first row we find the column with the active we want, active1 or 2
                        for cell in row:
                            if(
                                (isinstance(cell.value,unicode) or isinstance(cell.value,str)) and
                                (str(t*-1) in cell.value) and #for valid cell.value if it contains the number...
                                not cell_col_id #and cell_col_id is not found yet...
                            ):
                                cell_col_id= cell.column
                    else: #for second row and above
                        if(
                            isinstance(row[0].value,datetime) and
                            (
                                row[0].value.date() >= start_date
                                and
                                row[0].value.date() < end_date
                            )
                        ):
                            try:
                                # print ws[get_column_letter(cell_col_id)+str(row_idx)].internal_value
                                date_and_future_prices.append(
                                    (
                                        row[0].value.date(),
                                        float(ws[get_column_letter(cell_col_id)+str(row_idx)].value)
                                    )
                                )
                            except:
                                logger.warn("[__get_price_store_future] invalid cell '{}' for date {} ".format(get_column_letter(cell_col_id)+str(row_idx),row[0].value.date()))
                                raise
                        elif not row[0].value or row[0].value.date() < start_date: #if go beyond the startdate means not relevant anymore for the other rows
                            break;

            asset_name = u"{}_ACTIVE{}".format(asset_name,t*-1)
        elif t!=None and t >= 0:  #For T(x) future
            if get_mapping==False:
                for row in ws.rows:
                    row_idx = row[0].row
                    if row[0].row == 1: #for first row we find the column with the date we want. i.e. current date + 1 month
                        for cell in row:

                            if isinstance(cell.value,datetime):
                                cell_date = cell.value.replace(day=1).date()
                                cell_date_mth_str = cell_date.strftime("%Y%m")
                                for_date_str = (cell_date-relativedelta(months=t)).strftime("%Y%m")
                            if( #if the first row date in this cell is of interest, we add...
                                isinstance(cell.value,datetime) and
                                for_date_str in col_idx_by_mth
                            ):
                                col_idx_by_mth[for_date_str] = cell.column
                            # logger.info("--Process header {}".format(for_date_str))
                    else: #for second row and above
                        if(
                            isinstance(row[0].value,datetime) and
                            (
                                row[0].value.date() >= start_date
                                and
                                row[0].value.date() < end_date
                            )
                        ):
                            date_str = row[0].value.date().strftime("%Y%m")
                            # logger.info("*Processing {} vs start_date! {}, end_date {}".format(row[0].value,start_date,end_date))
                            try:
                                price = (ws[get_column_letter(col_idx_by_mth[date_str])+str(row_idx)].value)
                                # print price
                                price = float(price)
                                date_and_future_prices.append(
                                    (
                                        row[0].value.date(),
                                        price
                                    )
                                )
                                # logger.info("----PRICE: {}".format(price))
                            except:
                                if date_str in col_idx_by_mth:
                                    cell_id = str(col_idx_by_mth[date_str])+","+str(row_idx)
                                else:
                                    cell_id = "Future for t = {} not found in data.".format(t)
                                logger.warn("[__get_price_store_future] invalid cell '{}' for date {} ".format(cell_id,row[0].value.date()))
                                # raise
                                break #end if future price cannot be found..
                        elif not row[0].value or row[0].value.date() < start_date: #if go beyond the startdate means not relevant anymore for the other rows
                            break

            asset_name = u"{}_T{}".format(asset_name,t)
        else: #For just select column mode...

            cell_col_id = column_index_from_string(excel_col)
            for row in ws.rows:
                asset_sub_name=None
                row_idx = row[0].row
                # if row[0].row == 1: #for first row we find the column with the active we want, active1 or 2
                if not isinstance(row[0].value,datetime): #if not datetime means we coni
                    if not asset_name and not asset_sub_name: #get the sub naem only if asset name is not set
                        asset_sub_name=row[cell_col_id-1].value # minus one as row[idx] idx start from 0 where cell_col_id start from 1.
                else: #for second row and above
                    if get_mapping==True: #stop if just want to get mapping
                        break

                    if(
                        isinstance(row[0].value,datetime) and
                        (
                            row[0].value.date() >= start_date
                            and
                            row[0].value.date() < end_date
                        )
                    ):
                        #get the column needed
                        try:
                            date_and_future_prices.append(
                                (
                                    row[0].value.date(),
                                    float(ws[get_column_letter(cell_col_id)+str(row_idx)].value)
                                )
                            )
                        except:
                            logger.warn("[__get_price_store_future] invalid cell col'{}', row'{}' for date {}. Reading stop for this timeseries.".format(get_column_letter(cell_col_id),str(row_idx),row[0].value.date()))
                            # raise
                            break #terminate if the row has invalid price...
                    elif not row[0].value or row[0].value.date() < start_date: #if go beyond the startdate means not relevant anymore for the other rows
                        break;
            if not asset_name: #only if asset name is not set, we get subname
                asset_name = u"{}_{}".format(asset_name,asset_sub_name)

        df = pd.DataFrame(date_and_future_prices,columns=['Date',asset_name])
        return df
    except:
        logger.error("Error in _get_price_store_future")
        raise

    wb._archive.close()

def automan_cap_tag_bug_fix(tag="measure:usd"):
    '''
    For some strange reason, automan has a bug with upper or lower case tags.
    The tags are stored 'case-sensitively' but treated mostly 'case-insenstively'.
    Except in instance of updating tags which can take place via the web UI.
    For example, the tag, 'measure:USD', 'USD' is stored in cap. If you find a time series with
    this tag and you change 'USD' to 'usd' the entire 'measure:USD' tag will disappear.
    To workaround this bug, this function return the appropriate cap/non cap tag for known case sensitive tags
    '''
    replacement_mapping = {
        "measure:usd":"measure:USD",
        "measure:CNY":"measure:cny",
        "measure:index":"measure:Index",
        "measure:rate":"measure:Rate",
    }

    for key in replacement_mapping:
        tag = tag.replace(key,replacement_mapping[key])

    return tag

def load_data_pricestore(
                            t=3,
                            list_futures = [
                                {"name":'DCE.IOEACNY.Store',"measure":"CNY"},
                                {"name":'SHFE.RBTACNY.Store',"measure":"CNY"},
                                {"name":'SHFE.HRCTACNY.Store ',"measure":"CNY"},
                            ],
                            start_date=date(2014,9,1),
                            end_date=date(2016,10,1),
                            create_automan_ts = False, #set to true if you want to create in automan
                            copy_from_network = True,
                            excel_col=None,
                            get_mapping=False, #set to true if you only want to get mapping of sheet to ts_name
                            price_store_network_path = "//sgsing022m/sing8data/Metals Risk Team/Metals_risk_data/PriceStore/PriceStore.xlsb",
                            price_store_local_relative_path = "/rv/PriceStore.xlsm",
                            temp_data_path = 'C:/tmp_data',
                            tags="metals source:pricestore cmdty:metals_futures project:rv_pdt",
                            overwrite_tags=False,
                            check_automan=False #setting this to true will make an additional call to automan to fetch data.
                            ):
    '''
    Load data form price store into automan
    '''
    logger = logging.getLogger(__name__)

    price_store_local_path = "{}{}".format(temp_data_path,price_store_local_relative_path)

    if copy_from_network:
        xls_conversion_vbscript(price_store_network_path,price_store_local_path,temp_data_path,delete_vb=True) #comment if don't need to copy again from network.

    wb = openpyxl.load_workbook(price_store_local_path, read_only=True, data_only=True)

    mapping = {} # a dictionary containing mapping of time series name in automan to sheet
    for future in list_futures:
        logger.info("Processing {}...".format(future['name']))
        if t==None: #we loop through all columns for spot prices
            start_col = column_index_from_string(future["start_excel_col"])
            end_col = column_index_from_string(future["end_excel_col"])
        else:
            start_col = 1
            end_col = 2

        for i in xrange(start_col,end_col):
            if t==None:
                logger.info("  column {} of sheet {}".format(get_column_letter(i),future['name']))

            #Check current automan data if needed
            old_start_date = start_date
            if check_automan:
                result = _get_price_store_future( #get the price name in automan
                                        price_store_path=price_store_local_path,
                                        wb=wb,
                                        start_date=start_date,
                                        end_date=end_date,
                                        sheet_name = future['name'],
                                        t = t,
                                        excel_col=get_column_letter(i),
                                        get_mapping=True,
                                        asset_name=future['clean_name'],
                                        )
                ts_name = _clean_automan_ts_name(result.columns[1])
                ts_name = "PRICESTORE_"+ts_name

                current_ts = automan_read_by_name(ts_name,start_date=start_date-timedelta(366))
                if (len(current_ts.columns)==1): #meaning no such time series(ts) yet..
                    start_date=date(1970,1,1) #we load all date by setting start time to 1970
                    logger.info("--new TS ({}) detected. Overwriting start_date to 1970 to load everything.".format(ts_name))

            result = _get_price_store_future(
                                    price_store_path=price_store_local_path,
                                    wb=wb,
                                    start_date=start_date,
                                    end_date=end_date,
                                    sheet_name = future['name'],
                                    t = t,
                                    excel_col=get_column_letter(i),
                                    get_mapping=get_mapping,
                                    asset_name=future['clean_name'],
                                    )
            ts_name = _clean_automan_ts_name(result.columns[1])
            ts_name = "PRICESTORE_"+ts_name

            start_date=old_start_date #set it back in case strt date overwrite
            # print  result

            #create the time series in automan




            if t!=None and t >= 0:
                expiration_tag = "expiration:t{} type:future".format(t)
            elif t!=None and t < 0:
                expiration_tag = "expiration:active{} type:active".format(t*-1)
            else: #t==0
                expiration_tag = "expiration:spot type:spot"

            mapping[ts_name]=future # store the mapping of the ts_name to sheet name

            if create_automan_ts and get_mapping==False:
                if overwrite_tags:
                    tags_full="{} freq:undefined {} measure:{} unit:{}".format(tags,expiration_tag,future['measure'],future['measure'].lower())
                else:
                    tags_full="{} {} measure:{} unit:{}".format(tags,expiration_tag,future['measure'],future['measure'].lower())

                tags_full=automan_cap_tag_bug_fix(tags_full)
                # logger.info("=========="+tags_full)
                automan_create_ts(ts_name,ts_name+". Data from PriceStores.xlsb",tags=tags_full,period="sparse",overwrite_tags=overwrite_tags,logger=logger)
                #write to the time series
                result.columns = ["Date",ts_name]
                automan_write(result)

                logger.info("Processed {}! Written to Automan! TS name: {}".format(future['name'],ts_name))
            else:
                logger.info("Processed {}! Not written to Automan. create_automan_ts==False.".format(future['name']))


    return mapping








def load_data_from_wande_excel(
                            start_date=get_biz_date(-7), #default start date is 7 days ago
                            end_date=get_biz_date(1), #end_date is not included. Set to None to load all
                            file_path=u"C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\automan_wande\\wande_excel\\\wanda_eng.xlsx",
                            overwrite_tags=False, #set to true if you want to overwrite the tags in case where the time series exists
                            overwrite_description=False, #set to true if you want to overwrite the description
                            logger=None,
                            automan_period_type = "sparse",
                            create_backup=True, #set to true to create backup of the excel file...
                            parameter_file_path=u"C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\automan_wande\\wande_inputs.xlsx", #the excel file to write the tags data to
                            multiprocess=False, #if this is set to true, the parameter file will not be altered. Rather the rows to be added will be returned.
                            multiprocess_output_path = "c:\\tmp_data\\wande\\output",
                            tag_mapping=None,
                            thread_id=None
                            ):
    '''
    Read an excel that is generated from Wan de excel plugin. First col must be date followed by names from Wande.
    Function will create the time series name using the Wande name (it will be '_clean_automan_ts_name').
    At the end, the function will also create a sheet (or update) called 'tags' where each ts name is listed with it's tags to be updated.

    '''
    if not logger:
        logger = logging.getLogger(__name__)

    #create backup of the parameter excel file if needed.
    if create_backup:
        drive, path = os.path.splitdrive(parameter_file_path)
        path, filename = os.path.split(path)

        directory = os.path.dirname(parameter_file_path)
        directory_backup = directory+"\\backup"
        if not os.path.exists(directory_backup):
            os.makedirs(directory_backup)

        new_filename = filename.split(".")[0]+"_backup_"+datetime.now().strftime("%Y-%m-%d_%H:%M:%S.%f").replace(":","_").replace(".","_")+"."+filename.split(".")[1]
        backup_path =directory_backup+"\\"+new_filename
        copyfile(parameter_file_path, backup_path)


    logger.info(u"!!! ALERT: PLEASE ENSURE '{}' IS NOT OPENED! IT WILL BE UPDATED BY THIS SCRIPT. IF NOT SCRIPT WILL FAIL!!!".format(parameter_file_path))

    wb = openpyxl.load_workbook(file_path, read_only=True)
    if multiprocess:
        wb_parameter = openpyxl.load_workbook(parameter_file_path, read_only=True)
    else:
        wb_parameter = openpyxl.load_workbook(parameter_file_path, read_only=False)
    ws_list = wb.worksheets
    total_ws = len(ws_list)



    sys_start_time = datetime.now()


    #get list of name of ws
    ws_name_list=[]
    for ws in ws_list:
        ws_name_list.append(ws.title)

    if not multiprocess: #for single process mode
        tag_mapping = {}
        if "tags" in wb_parameter.get_sheet_names():
            ws_tags = wb_parameter["tags"]
        else:
            ws_tags = wb_parameter.create_sheet("tags")

        col_mapping = {
            "ts_name":"A",
            "clean_name":"B",
            "tags":"C",
            "wande_id":"D",
            "main_commodity_category":"E",
            "sub_commodity_category":"F",
            "main_data_type":"G",
            "sub_data_type":"H",
            "main_location":"I",
            "sub_location":"J",
            "automan_link":"K",
        }

        #add the header regardless whether it is new or old sheet
        for col in col_mapping:
            ws_tags['{}1'.format(col_mapping[col])]=col

        #process the mapping.
        tag_row_idx=1 #the current row pointer
        for row in ws_tags.rows:
            if tag_row_idx > 1:
                if row[0].value and row[1].value:
                    tag_mapping[row[0].value]= {
                                                "clean_name":row[(column_index_from_string(col_mapping["clean_name"])-1)].value,
                                                "row_idx":tag_row_idx,
                                                "tags":row[(column_index_from_string(col_mapping["tags"])-1)].value,
                                                "wande_id":row[(column_index_from_string(col_mapping["wande_id"])-1)].value,
                                                "main_commodity_category":row[(column_index_from_string(col_mapping["main_commodity_category"])-1)].value,
                                                "sub_commodity_category":row[(column_index_from_string(col_mapping["sub_commodity_category"])-1)].value,
                                                "main_data_type":row[(column_index_from_string(col_mapping["main_data_type"])-1)].value,
                                                "sub_data_type":row[(column_index_from_string(col_mapping["sub_data_type"])-1)].value,
                                                "main_location":row[(column_index_from_string(col_mapping["main_location"])-1)].value,
                                                "sub_location":row[(column_index_from_string(col_mapping["sub_location"])-1)].value,
                                                "automan_link":row[(column_index_from_string(col_mapping["automan_link"])-1)].value,
                                                }
            tag_row_idx+=1
        if len(tag_mapping)==0: #for cases where there isn't any records but has empty row.
            tag_row_idx=2

    else: #multiprocess, we will just store them into dictionary call returned_tag_mapping written to multiprocess_output_path
        returned_tag_mapping = {}
        pass


    #loop sheet
    ws_processed_count=0
    for ws in ws_list:
        price_processed_count=0
        ws_name = ws.title
        if ws_name != "tags": #we only process for non tags sheet

            logger.info(u"=Processing WS {} ({} out of {}) =".format(ws_name,str(ws_processed_count+1),str(total_ws)))

            num_col_in_ws = len(ws.columns)
            start_col = column_index_from_string("B")
            end_col = len(ws.columns)+1
            total_col = end_col-start_col
            last_row_excel_idx = ws.max_row #start from 1

            #Get what each row means first
            row_mapping = {}
            date_found = False
            date_start_row=1
            while(not date_found): #find the row with date
                cell_val = ws["A"+str(date_start_row)].value
                if isinstance(cell_val,datetime): #datetime found
                    date_found = True
                else: # datetime not found
                    if cell_val.lower() == "Name".lower() or cell_val==u"指标名称":
                        row_mapping['name']=date_start_row
                    elif cell_val.lower() == "Frequency".lower() or cell_val==u"频率":
                        row_mapping['frequency']=date_start_row
                    elif cell_val.lower() == "Unit".lower() or cell_val==u"单位":
                        row_mapping['unit']=date_start_row
                    elif cell_val.lower() == "ID".lower() or cell_val==u"指标ID":
                        row_mapping['id']=date_start_row
                    date_start_row+=1

            #loop each col
            for i in xrange(start_col,end_col):
                col_ltr = get_column_letter(i)
                #1) prepare the time series information:

                description = u""
                tags = u""
                name = u""
                wande_id = u""
                for key in row_mapping: #loop through all the meta data to prepare tag and description
                    val = ws["{0}{1}".format(col_ltr,str(row_mapping[key]))].value

                    if key == "frequency":
                        if val == u"日":
                            val="Daily"
                        elif val == u"周":
                            val="Weekly"
                        elif val == u"月":
                            val="Monthly"
                        elif val == u"年":
                            val="Yearly"

                    description+=u"{0}:{1}\n".format(key,val)

                    if key=="name": #don't add name into tags
                        continue

                    if key=="id":
                        wande_id=val

                    if key == "frequency":
                        tags +="{0} freq:{0} ".format(val.lower())
                    elif key == "unit":
                        tags +="measure:{0} unit:{0} ".format(_automan_tag_friendly(val))
                    else:
                        tags+="{}:{} ".format(key,_automan_tag_friendly(val))

                tags+="source:wind wind "
                ts_name = _clean_automan_ts_name(ws["{}{}".format(col_ltr,str(row_mapping['name']))].value)

                # print ts_name
                # print description
                # print tags

                #2) Create the time series
                    #If tags sheet has this time series, we ignore the generated tags above
                    #If tags sheet does not have this time series, we use generated tags above
                overwrite_tags = True
                overwrite_description = True
                if ts_name in tag_mapping: #if in tag_mapping already, meaning old TS that has already been loaded once
                    overwrite_tags = True
                    overwrite_description = False
                    tags = tag_mapping[ts_name]['tags']
                    clean_name = _clean_automan_ts_name(tag_mapping[ts_name]['clean_name'])
                    ts_row_idx_in_ws_tags = tag_mapping[ts_name]['row_idx']


                    tags=tags.strip()

                    cols_to_tags = [
                                        'main_commodity_category','sub_commodity_category',
                                        'main_data_type','sub_data_type',
                                        'main_location','sub_location'
                                   ]
                    for col in cols_to_tags:
                        if tag_mapping[ts_name][col]: #if there is a value for the col, then we add.
                            tags+=" {}:{}".format(col,_automan_tag_friendly(tag_mapping[ts_name][col]))

                    ts_name_old_name = ts_name #the old name is the one derive directly from Wande sheet
                    #rename if needed. and change ts_name to the clean name
                    if ts_name!=clean_name:
                        try:
                            automan_update_ts(name=ts_name,new_name=clean_name)
                        except:
                            logger.warn("==Renaming failed for changing {} to {}. Don't panick. It could be renamed before.".format(ts_name,clean_name))

                        if not multiprocess:
                            #update the sheet with the clean name
                            ws_tags["{}{}".format(col_mapping["clean_name"],ts_row_idx_in_ws_tags)]=clean_name
                            ws_tags["{}{}".format(col_mapping["automan_link"],ts_row_idx_in_ws_tags)]='=HYPERLINK("http://ot-metals.automan.cargill.com/data/{0}")'.format(quote(clean_name))

                        #change the ts_name to clean name
                        ts_name=clean_name

                    if multiprocess: #for multiprocess we update the returned_tag_mapping
                        returned_tag_mapping[ts_name_old_name]={
                            "ts_name":ts_name_old_name,
                            "clean_name":clean_name,
                            "tags":tags,
                            "wande_id":tag_mapping[ts_name_old_name]["wande_id"],
                            "main_commodity_category":_automan_tag_friendly(tag_mapping[ts_name_old_name]["main_commodity_category"]),
                            "sub_commodity_category":_automan_tag_friendly(tag_mapping[ts_name_old_name]["sub_commodity_category"]),
                            "main_data_type":_automan_tag_friendly(tag_mapping[ts_name_old_name]["main_data_type"]),
                            "sub_data_type":_automan_tag_friendly(tag_mapping[ts_name_old_name]["sub_data_type"]),
                            "main_location":_automan_tag_friendly(tag_mapping[ts_name_old_name]["main_location"]),
                            "sub_location":_automan_tag_friendly(tag_mapping[ts_name_old_name]["sub_location"]),
                            "automan_link":'=HYPERLINK("http://ot-metals.automan.cargill.com/data/{0}")'.format(quote(clean_name)),
                        }
                    logger.warn("==Mapping for ts name of '{}' already exists in 'tags' sheet of the workbook. Tags to be updated in automan used will be tags in 'tags sheet'. Tags: <{}>.".format(ts_name,tags))
                else:

                    if not multiprocess: #we write to excel if not multiprocess
                        #this is a new time series not in the tag, we create .
                        ws_tags["{}{}".format(col_mapping["ts_name"],tag_row_idx)]=ts_name
                        ws_tags["{}{}".format(col_mapping["clean_name"],tag_row_idx)]=ts_name
                        ws_tags["{}{}".format(col_mapping["tags"],tag_row_idx)]=tags
                        ws_tags["{}{}".format(col_mapping["wande_id"],tag_row_idx)]=wande_id

                        ws_tags["{}{}".format(col_mapping["main_commodity_category"],tag_row_idx)]=""
                        ws_tags["{}{}".format(col_mapping["sub_commodity_category"],tag_row_idx)]=""
                        ws_tags["{}{}".format(col_mapping["main_data_type"],tag_row_idx)]=""
                        ws_tags["{}{}".format(col_mapping["sub_data_type"],tag_row_idx)]=""
                        ws_tags["{}{}".format(col_mapping["main_location"],tag_row_idx)]=""
                        ws_tags["{}{}".format(col_mapping["sub_location"],tag_row_idx)]=""

                        ws_tags["{}{}".format(col_mapping["automan_link"],tag_row_idx)]='=HYPERLINK("http://ot-metals.automan.cargill.com/data/{0}")'.format(quote(ts_name))
                        #add to the tag mapping too...
                        tag_mapping[ts_name]= {"row_idx":tag_row_idx,"tags":tags}
                        #move to new row
                        tag_row_idx+=1
                    else: #multiprocess we update returned_tag_mapping
                        returned_tag_mapping[ts_name]={
                            "ts_name":ts_name,
                            "clean_name":ts_name,
                            "tags":tags,
                            "wande_id":wande_id,
                            "main_commodity_category":"",
                            "sub_commodity_category":"",
                            "main_data_type":"",
                            "sub_data_type":"",
                            "main_location":"",
                            "sub_location":"",
                            "automan_link":'=HYPERLINK("http://ot-metals.automan.cargill.com/data/{0}")'.format(quote(ts_name)),
                        }


                    #change the time series start date so we load all data...
                    start_date=date(1900,1,1)

                #FINALLY Create the time series in automan
                tags_full="{0} freq:undefined period:{1}".format(tags,automan_period_type)
                logger.info("==Processing {0}".format(ts_name))

                logger.info("===a) Creating or updating time series {0} with tags <{3}>. overwrite_tags={1}, overwrite_description={2},automan_period_type={4}".\
                        format(
                                ts_name,str(overwrite_tags),str(overwrite_description)
                                ,tags,automan_period_type
                              )
                        )
                automan_create_ts(name=ts_name,description=description,tags=tags,period=automan_period_type,overwrite_tags=overwrite_tags,overwrite_description=overwrite_description,logger=logger)
                logger.info("===>a) Created or updated time series {0}!".format(ts_name))

                #3) By now tags table updated and automan should have the time series created
                logger.info("===b) Loading time series {0} into automan".format(ts_name))
                df = load_single_time_series_automan(
                    mode="excel",
                    opts={"path":file_path,"ws":ws_name,
                          "date_col_letter_idx":"A",
                          "time_series_letter_idx":col_ltr,
                          "start_row_excel_idx":date_start_row},
                    time_series_name=ts_name,
                    start_date=start_date,
                    end_date = end_date
                )
                price_processed_count+=1
                price_percentage_completed = float(price_processed_count)/float(total_col)*100.0
                logger.info("===>b) Loaded time series {0} into automan! Total Elapsed Time: {1} | {2} out of {3} prices for sheet '{4}' [{5:.2f}%]".\
                            format(
                                ts_name,
                                _get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now()),
                                price_processed_count,
                                total_col,
                                ws_name,
                                price_percentage_completed
                                )
                            )
                logger.info("==>Processed {0}!".format(ts_name))


            ws_processed_count+=1
            ws_percentage_completed = float(ws_processed_count)/float(total_ws)*100.0
            logger.info(
                        u"=Finish Processing WS:{0}= Total Elapsed Time: {1} | {2} out of {3} sheets completed [{4:.2f}%]"\
                        .format(
                                    ws_name,_get_human_readable_time_diff(start_time=sys_start_time,end_time=datetime.now()),
                                    ws_processed_count,
                                    str(total_ws),
                                    ws_percentage_completed
                                )
                        )


        #save the workbook
        if not multiprocess:
            logger.info("Saving {} with update tags sheet...".format(parameter_file_path))
            wb_parameter.save(parameter_file_path)
            logger.info("Saved {}!".format(parameter_file_path))
        else:
            output_file_path = "{}/{}_output.p".format(multiprocess_output_path,thread_id)
            logger.info("Saving output dictionary mapping to {}...".format(output_file_path))
            #save the output to dictionary and save to disk
            pickle.dump(returned_tag_mapping,open(output_file_path,"wb"))
            logger.info("Saved output dictionary mapping to {}!".format(output_file_path))

    logger.info("All numbers loaded to Automan!")
    return returned_tag_mapping










def load_data_from_mysteel_excel(
                            start_date=get_biz_date(-7), #by default load last 1 week data. Set to None to load all
                            end_date=get_biz_date(1), #end_date is not included. Set to None to load all
                            file_path=u"C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\Rick's Daily Prices\\Daily prices_mapping.xlsx",
                            overwrite_tags=False, #set to true if you want to overwrite the tags in case where the time series exists
                            overwrite_description=False, #set to true if you want to overwrite the description
                            start_row_excel_idx = 11, #by default we start reading from row 11 onwards
                            mapping={
                                        #key is the sheet name. row start from 1
                                        "Rebar prices": {"series_name_row":1,"tags_row":2,"description_row":3,"start_data_col_excel":"B","end_data_col_excel":"AB","ignore_cols":["C","D"]}
                                    },
                            logger=None
                            ):
    if not logger:
        logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    df_dict={}
    for key, map_ts in mapping.iteritems():
        ws = wb[key]
        ws_name = key
        logger.info(u"=Processing WS {}=".format(ws_name))

        #loop each col
        start_col = column_index_from_string(map_ts["start_data_col_excel"])
        end_col = column_index_from_string(map_ts["end_data_col_excel"])
        for i in xrange(start_col,end_col):
            col_ltr = get_column_letter(i)
            map_ts['data_col_excel'] = col_ltr

            if "start_row_excel_idx" in map_ts:
                start_row_excel_idx=map_ts["start_row_excel_idx"]

            if col_ltr not in map_ts['ignore_cols']:#process only if it is not in the ignore list

                series_name = ws[map_ts['data_col_excel']+str(map_ts['series_name_row'])].value
                series_tags = ws[map_ts['data_col_excel']+str(map_ts['tags_row'])].value

                series_description = ws[map_ts['data_col_excel']+str(map_ts['description_row'])].value
                time_series_letter_idx = map_ts['data_col_excel']
                start_row_excel_idx = start_row_excel_idx

                logger.info(u"  Processing WS:{} Col:{} Series:{}".format(ws_name,col_ltr,series_name))


                #1. First create the time series
                tags_full="{} freq:undefined period:sparse".format(series_tags)
                # logger.info("===tags_full:{}".format(tags_full))
                automan_create_ts(name=series_name,description=series_name+" "+series_description,tags=tags_full,period="sparse",overwrite_tags=overwrite_tags,overwrite_description=overwrite_description)

                #2. Load the time series into automan
                df = load_single_time_series_automan(
                    mode="excel",
                    opts={"path":file_path,"ws":ws_name,"date_col_letter_idx":"A","time_series_letter_idx":time_series_letter_idx,"start_row_excel_idx":start_row_excel_idx},
                    time_series_name=series_name,
                    start_date=start_date,
                    end_date = end_date
                )
                df_dict[series_name]=df

        logger.info(u"=Finish Processing WS:{}=".format(ws_name))
    return df_dict


def _clean_automan_ts_name(name):
    '''
    make name ok for automan to take.
    Maximum 100 characters
    '''
    string_limit=100
    replace_array = [
        (u'\u3000'," "),#remove irregular space
        (".","_"),
        ("%","pc"),
        # ("/","_"),
        # ("\\","_"),
        ("&","_AND_"),
        (">","_MORE_THAN_"),
        ("<","_LESS_THAN_"),
        (":","_"),
    ]
    clean_str=unicode(name)
    for replace_tuple in replace_array:
        clean_str = clean_str.replace(replace_tuple[0],replace_tuple[1])



    return clean_str.upper()[:string_limit]


def update_paper_price(
                        start_date=date.today(),
                        temp_data_path = 'C:/tmp_data',
                        paper_price_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Tableau/Price Analysis/paper price.xlsx",
                        paper_price_xlsx_local_relative_path = "/paper_price/paper price.xlsx",
                        price_store_network_path = "//sgsing022m/sing8data/Metals Risk Team/Metals_risk_data/PriceStore/PriceStore.xlsb",
                        price_store_local_relative_path = "/paper_price/PriceStore.xlsm",
                        shfe_rebar_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=RB&hy=RB0&breed=RB0&type=inner",
                        dce_coke_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=dce&pz=J&hy=J0&breed=J0&type=inner",
                        dce_io_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=dce&pz=I&hy=I0&breed=I0&type=inner",
                        copy_paper_price_from_network = True,
                        save_paper_price_to_network = False,
                        backup=True,
                        the_date = None,
                        suppress_exception=False
                      ):
    '''
    REQUIRED SCRAPPING:
        _get_sina_finance_future_price
        PriceStore.xlsb file
    Summary: Function will update paper price.xlsx in //sgsing022m/sing8data/ETIA/DRY/FERROUS/Tableau/Price Analysis/paper price.xlsx
             for the Tableau dashboard
    Author: Jax
    Date created: 14 Jul 2016
    Date last updated: 19 Jul 2016
    Precondition:
        Required: N/A
        Optional:
            start_date:
                The date of processing. Note that it will be the previous business day that will be process.I.E.,
                Dfeault .start_date=date.today()
            the_date:
                if the_date is set, the start_date will not be taken into account and the date to process will be the_date and not the previous biz day from start_date
            temp_data_path:
                the temporary data path in your local machines. Default temp_data_path = 'C:/Users/j256377/Desktop/Metals/AutoIntern/tmp_data',
            paper_price_xlsx_network_path:
                the network path with the paper_price.xlsx. Default paper_price_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Tableau/Price Analysis/paper price.xlsx",
            paper_price_xlsx_local_relative_path:
                Default paper_price_xlsx_local_relative_path = "/paper_price/paper price.xlsx",
            price_store_network_path:
                The price store on network drive.
                Default price_store_network_path = "//sgsing022m/sing8data/Metals Risk Team/Metals_risk_data/PriceStore/PriceStore.xlsb"
            price_store_local_relative_path:
                Default price_store_local_relative_path = "/paper_price/PriceStore.xlsm",
            shfe_rebar_nodate_url:
                The SH rebar sina webpage without the date.
                Default shfe_rebar_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=RB&hy=RB0&breed=RB0&type=inner",
            dce_coke_nodate_url:
                The DCE coke sina webpage without the date.
                Default dce_coke_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=dce&pz=J&hy=J0&breed=J0&type=inner",
            dce_io_nodate_url:
                The DCE IO sina webpage without the date.
                Default dce_io_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=dce&pz=I&hy=I0&breed=I0&type=inner",
            copy_paper_price_from_network:
                Set to true if you want to copy the latest paper_price.xlsx from the network.
                Default copy_paper_price_from_network = True,
            save_paper_price_to_network:
                Set to true if you want to save the paper_price.xlsx back to network drive.
                Default save_paper_price_to_network = False
            backup:
                Set to true if you want to have a copy of the network file in local as backup before processing. Default is True

    Postcondition:
    '''
    logger = logging.getLogger(__name__)

    #try to update paper price, return error if not
    try:
        #0. Prepare data...
        if not the_date:
            today = start_date
            yesterday = get_biz_date(-1,today)
        else:
            yesterday = the_date
            today = get_biz_date(1,yesterday)
        day_before_yesterday = get_biz_date(-1,yesterday)
        today_str = get_date_str_frm_date(today)#get_date_str()
        yesterday_str = get_date_str_frm_date(yesterday)#get_date_str(-1)


        logger.info("Processing for today {} looking at the previous business day {} numbers".format(today.strftime("%Y-%m-%d (%a)"),yesterday.strftime("%Y-%m-%d (%a)")))

        paper_price_xlsx_local_path = "{}{}".format(temp_data_path,paper_price_xlsx_local_relative_path)
        price_store_local_path = "{}{}".format(temp_data_path,price_store_local_relative_path)
        # xlsb_to_xls_path = "C:\\Users\\j256377\\Desktop\\Metals\\AutoIntern\\xlsb_to_xlsx.vbs"

        # shfe_rebar_url = "{}&start={}&end={}".format(shfe_rebar_nodate_url,yesterday_str,yesterday_str)
        # dce_coke_url = "{}&start={}&end={}".format(dce_coke_nodate_url,yesterday_str,yesterday_str)
        # dce_io_url = "{}&start={}&end={}".format(dce_io_nodate_url,yesterday_str,yesterday_str)


        #1. Get Prices form MySteel
            #a) SHFE Rebar
        try:
            # r  = requests.get(shfe_rebar_url)
            # data = r.text
            # soup = BeautifulSoup(data)
            # tr_history = soup.find("tr", {"class" : "tr_2"})
            # shfe_rebar_price = float(tr_history.find_all('div')[1].contents[0])
            shfe_rebar_price = _get_sina_finance_future_price(url=shfe_rebar_nodate_url,the_date=yesterday,suppress_exception=suppress_exception)
        except:
            logger.error("Error in scraping SHFE Rebar at {}".format(shfe_rebar_nodate_url))
            raise

            #b) DCE Coke
        try:
            # r  = requests.get(dce_coke_url)
            # data = r.text
            # soup = BeautifulSoup(data)
            # tr_history = soup.find("tr", {"class" : "tr_2"})
            # dce_coke_price = float(tr_history.find_all('div')[1].contents[0])
            dce_coke_price = _get_sina_finance_future_price(url=dce_coke_nodate_url,the_date=yesterday,suppress_exception=suppress_exception)
        except:
            logger.error("Error in scraping DCE Coke at {}".format(dce_coke_url))
            raise

            #c) DCE Iron Ore
        try:
            # r  = requests.get(dce_io_url)
            # data = r.text
            # soup = BeautifulSoup(data)
            # tr_history = soup.find("tr", {"class" : "tr_2"})
            # dce_io_price = float(tr_history.find_all('div')[1].contents[0])
            dce_io_price = _get_sina_finance_future_price(url=dce_io_nodate_url,the_date=yesterday,suppress_exception=suppress_exception)
        except:
            logger.error("Error in scraping DCE IO at {}".format(dce_coke_url))
            raise

        logger.info("shfe rebar: {}, dce coke: {}, dce io: {} (sina.com.cn)".format(shfe_rebar_price, dce_coke_price, dce_io_price))





        #2. Get SGX IO and SSEC HRC Prices from PriceStore
        xls_conversion_vbscript(price_store_network_path,price_store_local_path,temp_data_path,delete_vb=True)

        wb = openpyxl.load_workbook(price_store_local_path, read_only=True)
            #d) Get SGX IO 2nd Mth price
        column_idx = None
        row_idx = None
        try:
            ws = wb['SGX.IO.Store']
            row_found=False
            for row in ws.rows:
                if not row_found:
                    if row[0].row == 1: #for first row we find the column with the date we want. i.e. current date + 1 month
                        for cell in row:
                            if(
                                isinstance(cell.value,datetime) and
                                cell.value.replace(day=1).date() == (yesterday+relativedelta(months=1)).replace(day=1)
                            ):
                                column_idx = cell.column
                                # print(cell.value.date())
                                # print(cell.column)
                    else: #for second row and above
                        if isinstance(row[0].value,datetime) and row[0].value.date() == yesterday:
                            row_idx = row[0].row
                            # print(row_idx)
                            row_found = True
                            break
                else: #if row_found we end this loop
                    break

            sgx_io_2_mth_price = ws[get_column_letter(column_idx)+str(row_idx)].value
        except:
            logger.error("Error in reading PriceStore.xlsm for SGX.IO 2nd Month at {}".format(price_store_local_path))
            raise


            #e) Get SSEC HRC price
        try:
            ws = wb['SSEC.HRC.Store']
            row_found=False
            column_idx = None
            row_idx = None
            for row in ws.rows:
                if not row_found:
                    if row[0].row == 1: #for first row we find the column with the date we want. i.e. current date + 1 month
                        for cell in row:
                            if(
                                isinstance(cell.value,datetime) and
                                cell.value.replace(day=1).date() == (yesterday+relativedelta(months=1)).replace(day=1)
                            ):
                                column_idx = cell.column
                    else: #for second row and above
                        if isinstance(row[0].value,datetime) and row[0].value.date() == yesterday:
                            row_idx = row[0].row
                            row_found = True
                            break
                else: #if row_found we end this loop
                    break

            ssec_hrc_price = float(ws[get_column_letter(column_idx)+str(row_idx)].value)
        except:
            logger.error("Error in reading PriceStore.xlsm for SSEC.HRC.Store at {}".format(price_store_local_path))
            raise

        logger.info("sgx io 2 mth: {}, ssec hrc: {}".format(sgx_io_2_mth_price,ssec_hrc_price))

        #close the workbook
        # wb._archive.close()



        #3. Write to "paper price.xlsx"
        #copy from network to tmp...
        directory = os.path.dirname(paper_price_xlsx_local_path)
        if not os.path.exists(directory):
            os.makedirs(directory)
        if copy_paper_price_from_network:
            shutil.copy(paper_price_xlsx_network_path ,paper_price_xlsx_local_path)


        wb = openpyxl.load_workbook(paper_price_xlsx_local_path)#,keep_vba=True)

        try:
            #a) price sheet
            ws = wb['price']
            date_col_id = None
            shfe_rebar_col_id = None
            sgx_io_2_mth_col_id = None
            dce_coke_col_id = None
            ssec_hrc_col_id = None
            row_to_add_row_id = None
            for row in ws.rows:
                if not row_to_add_row_id:
                    if row[0].row == 1: #for first row we find the column id we want
                        for cell in row:
                            if cell.value:  #for none empty cell
                                if cell.value.strip() == "Date":
                                    date_col_id = cell.column
                                elif cell.value.strip() == "SHFE Rebar active":
                                    shfe_rebar_col_id = cell.column
                                elif cell.value.strip() == "SGX IO 2nd Month":
                                    sgx_io_2_mth_col_id = cell.column
                                elif cell.value.strip() == "DCE Coke active":
                                    dce_coke_col_id = cell.column
                                elif cell.value.strip() == "SSEC HRC Store":
                                    ssec_hrc_col_id = cell.column
                    else: #for second row and above, first column must be date...
                        if isinstance(row[0].value,datetime) and row[0].value.date() == day_before_yesterday:
                            row_to_add_row_id = row[0].row + 1
                            break
                else: #if row_found we end this loop
                    break

            logger.info("Found row to add at {}. date_col_id at {}; shfe_rebar_col_id at {}; sgx_io_2_mth_col_id at {};  dce_coke_col_id at {};  ssec_hrc_col_id at {};".format(row_to_add_row_id,date_col_id,shfe_rebar_col_id,sgx_io_2_mth_col_id,dce_coke_col_id,ssec_hrc_col_id))


            #write to the sheet
            date_cell = '{}{}'.format(date_col_id,str(row_to_add_row_id))
            ws[date_cell] = yesterday
            ws[date_cell].number_format = ws['{}{}'.format(date_col_id,str(row_to_add_row_id-1))].number_format

            shfe_rebar_cell = '{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id));
            ws[shfe_rebar_cell].alignment = ws['{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id-1))].alignment
            ws[shfe_rebar_cell] = shfe_rebar_price

            sgx_io_2_mth_cell = '{}{}'.format(sgx_io_2_mth_col_id,str(row_to_add_row_id))
            ws[sgx_io_2_mth_cell].alignment = ws['{}{}'.format(sgx_io_2_mth_col_id,str(row_to_add_row_id-1))].alignment
            ws[sgx_io_2_mth_cell] = sgx_io_2_mth_price

            dce_coke_cell = '{}{}'.format(dce_coke_col_id,str(row_to_add_row_id))
            ws[dce_coke_cell].alignment = ws['{}{}'.format(dce_coke_col_id,str(row_to_add_row_id-1))].alignment
            ws[dce_coke_cell] = dce_coke_price

            ssec_hrc_cell = '{}{}'.format(ssec_hrc_col_id,str(row_to_add_row_id))
            ws[ssec_hrc_cell].alignment = ws['{}{}'.format(ssec_hrc_col_id,str(row_to_add_row_id-1))].alignment
            ws[ssec_hrc_cell] = ssec_hrc_price

            #b) data sheet
            ws = wb['data']
            date_col_id = "A"
            shfe_rebar_col_id = "B"
            sgx_io_date_col_id = "C"
            sgx_io_2_mth_col_id = "D"
            dce_coke_date_col_id = "F"
            dce_coke_col_id= "G"
            dce_io_col_id = "H"
            row_to_add_row_id = 3

            #insert row only if the date is not created yesterday
            if ws["{}{}".format(date_col_id,str(row_to_add_row_id))].value.date() != yesterday:
                insert_rows(ws=ws,row_idx=row_to_add_row_id-1,cnt=1)

            #add the value...
            date_cell = '{}{}'.format(date_col_id,str(row_to_add_row_id))
            ws[date_cell] = yesterday
            ws[date_cell].number_format = ws['{}{}'.format(date_col_id,str(row_to_add_row_id+1))].number_format

            shfe_rebar_cell = '{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id));
            ws[shfe_rebar_cell] = shfe_rebar_price
            ws[shfe_rebar_cell].alignment = ws['{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id+1))].alignment

            sgx_io_2_mth_cell = '{}{}'.format(sgx_io_2_mth_col_id,str(row_to_add_row_id))
            ws[sgx_io_2_mth_cell] = sgx_io_2_mth_price
            ws[sgx_io_2_mth_cell].alignment = ws['{}{}'.format(sgx_io_2_mth_col_id,str(row_to_add_row_id+1))].alignment

            dce_coke_cell = '{}{}'.format(dce_coke_col_id,str(row_to_add_row_id))
            ws[dce_coke_cell] = dce_coke_price
            ws[dce_coke_cell].alignment = ws['{}{}'.format(dce_coke_col_id,str(row_to_add_row_id+1))].alignment

            dce_io_cell = '{}{}'.format(dce_io_col_id,str(row_to_add_row_id))
            ws[dce_io_cell] = dce_io_price
            ws[dce_io_cell].alignment = ws['{}{}'.format(dce_io_col_id,str(row_to_add_row_id+1))].alignment


            sgx_io_date_cell = '{}{}'.format(sgx_io_date_col_id,str(row_to_add_row_id))
            ws[sgx_io_date_cell] = "={}".format(date_cell)
            ws[sgx_io_date_cell].number_format = ws[date_cell].number_format

            dce_coke_date_cell = '{}{}'.format(dce_coke_date_col_id,str(row_to_add_row_id))
            ws[dce_coke_date_cell] = "={}".format(sgx_io_date_cell)
            ws[dce_coke_date_cell].number_format = ws[sgx_io_date_cell].number_format



            wb.save(paper_price_xlsx_local_path)

            #Backup network copy first
            if backup:
                shutil.copy(paper_price_xlsx_network_path,paper_price_xlsx_local_path.replace(".xlsx","-{}.xlsx".format(today_str)))

            #copy to network path....
            if save_paper_price_to_network:
                shutil.copy(paper_price_xlsx_local_path,paper_price_xlsx_network_path)


            logger.info("Found row to add at {}. date_col_id at {}; shfe_rebar_col_id at {}; sgx_io_2_mth_col_id at {};  dce_coke_col_id at {};  ssec_hrc_col_id at {};".format(row_to_add_row_id,date_col_id,shfe_rebar_col_id,sgx_io_2_mth_col_id,dce_coke_col_id,ssec_hrc_col_id))

        except:
            logger.error("Error in writing to paper price.xlsx at path. Could be data for {} not in PriceStore".format(yesterday))
            raise

    except: #try reading and writing data... if fail...
        logger.error("update_paper_price() failed! Check whether: 1)\n{},\n2){}\n3){}\n4){}\nhas the data for this day {}".format(price_store_local_path,shfe_rebar_nodate_url,dce_coke_nodate_url,dce_io_nodate_url,yesterday_str))
        raise







def update_prices_xlsx(
                        start_date=date.today(),
                        temp_data_path = 'C:/tmp_data',
                        price_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx",
                        price_xlsx_network_tmp_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data_tmp.xlsx",
                        price_xlsx_local_relative_path = "/prices/Prices_data.xlsx",
                        shfe_rebar_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=RB&hy=RB0&breed=RB0&type=inner",
                        shfe_hrc_nodate_url = "http://vip.stock.finance.sina.com.cn/q/view/vFutures_History.php?jys=shfe&pz=HC&hy=HC0&breed=HC0&type=inner",
                        copy_price_from_network = True,
                        save_price_to_network = False,
                        backup=True,
                        the_date = None,
                        suppress_exception = False
                        ):
    '''
    S:\ETIA\DRY\FERROUS\Steel\Wei\Prices.xlsx
    REQUIRED SCRAPPING:
        _get_mysteel_price
        _get_sina_finance_future_price
        _get_sina_finance_fx
        _get_thesteelindex_price
        _get_feigang_price
        _get_steelbb_price
    Summary: Function will update Price_data.xlsx in //sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/Wei/Prices_data.xlsx.
            Note that Prices_data.slsx only have the figures and no graph. Prices_link.xlsx will have the graphs.
            Prices_link.xlsx reads data from Prices_data.xlsx.
    Author: Jax
    Date created: 14 Jul 2016
    Date last updated: 19 Jul 2016
    Precondition:
        Required: N/A
        Optional:
            start_date:
                The date of processing. Note that it will be the previous business day that will be process.I.E.,
                Dfeault .start_date=date.today()
            the_date:
                if the_date is set, the start_date will not be taken into account and the date to process will be the_date and not the previous biz day from start_date
            temp_data_path:
                the temporary data path in your local machines. Default temp_data_path = 'C:/Users/j256377/Desktop/Metals/AutoIntern/tmp_data',
    '''

    try:
        #PART 1: PRE-PROCESS
        logger = logging.getLogger(__name__)
        if not the_date:
            today = start_date
            yesterday = get_biz_date(-1,today)
        else:
            yesterday = the_date
            today = get_biz_date(1,yesterday)
        day_before_yesterday = get_biz_date(-1,yesterday)
        today_str = get_date_str_frm_date(today)#get_date_str()
        yesterday_str = get_date_str_frm_date(yesterday)#get_date_str(-1)


        logger.info("Processing for today {} looking at the previous business day {} numbers".format(today.strftime("%Y-%m-%d (%a)"),yesterday.strftime("%Y-%m-%d (%a)")))

        #Copy Prices.xlsx to local.
        price_xlsx_local_path = "{}{}".format(temp_data_path,price_xlsx_local_relative_path)
        if copy_price_from_network:
            _copy(price_xlsx_network_path,price_xlsx_local_path)

        #load the excel file for writing
        wb = openpyxl.load_workbook(price_xlsx_local_path)



        #PART II: Update China Prices sheet
        try:
            #a) Get Data
            #login to mysteel to get a session
            session_requests = _get_login_session(type="mysteel")
            #get mysteel price
            # ws = wb['ChinaPrices']
            # prices_mysteel = [
            #     {"cell_col_id":"D","value":None,"name":"Tangshan billet","payload":{"url":"http://list1.mysteel.com/article/p-3571-------------1.html","mode":"furnace_raw"}},
            #     {"cell_col_id":"F","value":None,"name":"Tianjing rebar 16-25mm Hua Dong","payload":{"url":"http://list1.mysteel.com/market/p-228-----010101-0-0104-------1.html","mode":"construction","price_name":u"天津市场建筑钢材价格行情","product":u"螺纹钢","spec":u"Ф16-25","origin":u"东华钢铁"}},
            #     {"cell_col_id":"AO","value":None,"name":"Tianjing HRC 4.75","payload":{"url":"http://list1.mysteel.com/price/p-10058--010103--1.html","mode":"hrc","area":u"天津","spec":u"4.75热轧板卷"}},
            #     {"cell_col_id":"I","value":None,"name":"Shanghai rebar 16-22mm","payload":{"url":"http://list1.mysteel.com/market/p-228-15278-----0--------1.html","mode":"construction",u"price_name":"上海市场建筑钢材价格行情","product":u"螺纹钢","spec":u"Φ16-22","origin":u"萍钢"}},
            #     {"cell_col_id":"AD","value":None,"name":"Shanghai CRC 1mm","payload":{"url":"http://list1.mysteel.com/price/p-10053--010104--1.html","mode":"crc","area":"上海","spec":"1.0mm"}},
            # ]

            ws = wb['ChinaPrices']
            prices_mysteel = [
                {"cell_col_id":"B","value":None,"name":"Tangshan billet","payload":{"url":"http://list1.mysteel.com/article/p-3571-------------1.html","mode":"furnace_raw"}},
                {"cell_col_id":"C","value":None,"name":"Tianjing rebar 16-25mm Hua Dong","payload":{"url":"http://list1.mysteel.com/market/p-228-----010101-0-0104-------1.html","mode":"construction","price_name":u"天津市场建筑钢材价格行情","product":u"螺纹钢","spec":u"Ф16-25","origin":u"东华钢铁"}},
                {"cell_col_id":"D","value":None,"name":"Shanghai rebar 16-22mm","payload":{"url":"http://list1.mysteel.com/market/p-228-15278-----0--------1.html","mode":"construction",u"price_name":"上海市场建筑钢材价格行情","product":u"螺纹钢","spec":u"Φ16-22","origin":u"萍钢"}},
                # {"cell_col_id":"D","value":None,"name":"Shanghai rebar 16-22mm","payload":{"url":"http://list1.mysteel.com/market/p-228-15278-----0--------1.html","mode":"construction",u"price_name":"上海市场建筑钢材价格行情","product":u"螺纹钢","spec":u"Φ16-25","origin":u"济钢"}},
                {"cell_col_id":"E","value":None,"name":"Shanghai CRC 1mm","payload":{"url":"http://list1.mysteel.com/price/p-10053--010104--1.html","mode":"crc","area":"上海","spec":"1.0mm"}},
                {"cell_col_id":"F","value":None,"name":"Tianjing HRC 4.75","payload":{"url":"http://list1.mysteel.com/price/p-10058--010103--1.html","mode":"hrc","area":u"天津","spec":u"4.75热轧板卷"}},
            ]
            for price in prices_mysteel:
                price['value'] = _get_mysteel_price(**dict({"suppress_exception":suppress_exception,"the_date":yesterday,"session_requests":session_requests}.items()+price['payload'].items()))
                logger.info("{} price: {} (www.mysteel.com)".format(price['name'],str(price['value'])))


            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
            row_to_add_row_id = result['row_to_add_row_id']
            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            for price in prices_mysteel:
                cell = '{}{}'.format(price["cell_col_id"],str(row_to_add_row_id));
                ws[cell] = price['value']
                ws[cell].number_format = ws['{}{}'.format(price["cell_col_id"],str(row_to_add_row_id-1))].number_format


        except:
            logger.error("II) Error in updating ChinaPrices get the URL of mysteel")
            raise

        #PART III: Update [SHFE Data]
        try:
                #a) Fetch data from sina
            shfe_rebar_price = _get_sina_finance_future_price(url=shfe_rebar_nodate_url,the_date=yesterday,suppress_exception=suppress_exception)
            logger.info("shfe_rebar_price: {} (sina.com.cn)".format(str(shfe_rebar_price)))
            shfe_hrc_price = _get_sina_finance_future_price(url=shfe_hrc_nodate_url,the_date=yesterday,suppress_exception=suppress_exception)
            logger.info("shfe_hrc_price: {} (sina.com.cn)".format(str(shfe_hrc_price)))
                #b) Update the worksheet
            # ws = wb['SHFE Data']
            # result = _find_row_to_add_by_date(ws=ws,date_col_idx=2,the_date=yesterday)
            # row_to_add_row_id = result['row_to_add_row_id']
            #
            # _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="C",the_date=yesterday)
            #
            # shfe_rebar_col_id = "D"
            # shfe_rebar_cell = '{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id));
            # ws[shfe_rebar_cell] = shfe_rebar_price
            #
            # date_col_id = "X"
            # date_cell = "{}{}".format(date_col_id,str(row_to_add_row_id))
            # ws[date_cell]=yesterday
            # ws[date_cell].number_format = ws['{}{}'.format(date_col_id,str(row_to_add_row_id-1))].number_format
            #
            # shfe_hrc_col_id = "Y"
            # shfe_hrc_cell = '{}{}'.format(shfe_hrc_col_id,str(row_to_add_row_id));
            # ws[shfe_hrc_cell] = shfe_hrc_price

            ws = wb['SHFE']
            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
            row_to_add_row_id = result['row_to_add_row_id']

            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            shfe_rebar_col_id = "B"
            shfe_rebar_cell = '{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id));
            ws[shfe_rebar_cell] = shfe_rebar_price
            ws[shfe_rebar_cell].number_format = ws['{}{}'.format(shfe_rebar_col_id,str(row_to_add_row_id-1))].number_format

            shfe_hrc_col_id = "C"
            shfe_hrc_cell = '{}{}'.format(shfe_hrc_col_id,str(row_to_add_row_id));
            ws[shfe_hrc_cell] = shfe_hrc_price
            ws[shfe_hrc_cell].number_format = ws['{}{}'.format(shfe_hrc_col_id,str(row_to_add_row_id-1))].number_format




        except:
            logger.error("III) Error in updating SHFE Data sheet. Check the URL '{}','{}'".format(shfe_rebar_nodate_url,shfe_hrc_nodate_url))
            raise



        #PART IV: Update [FXRates]
        try:
            #a) Get data
            usdcny = _get_sina_finance_fx(the_date=yesterday,cur_from="usd",cur_to="cny",suppress_exception=suppress_exception)
            logger.info("USD to CNY: {} (sina.com.cn)".format(str(usdcny)))
            usdtry = _get_sina_finance_fx(the_date=yesterday,cur_from="usd",cur_to="try",suppress_exception=suppress_exception)
            logger.info("USD to TRY: {} (sina.com.cn)".format(str(usdtry)))

            #b) Update worksheet
            ws = wb['FXRates']

            # result = _find_row_to_add_by_date(ws=ws,date_col_idx=2,the_date=yesterday)
            # row_to_add_row_id = result['row_to_add_row_id']
            #
            # _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="C",the_date=yesterday)
            #
            # usdcny_col_id = "D"
            # usdcny_cell = '{}{}'.format(usdcny_col_id,str(row_to_add_row_id));
            # ws[usdcny_cell] = usdcny
            #
            # usdtry_col_id = "L"
            # usdtry_cell = '{}{}'.format(usdtry_col_id,str(row_to_add_row_id));
            # ws[usdtry_cell] = usdtry

            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
            row_to_add_row_id = result['row_to_add_row_id']

            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            usdcny_col_id = "B"
            usdcny_cell = '{}{}'.format(usdcny_col_id,str(row_to_add_row_id));
            ws[usdcny_cell] = usdcny

            usdtry_col_id = "C"
            usdtry_cell = '{}{}'.format(usdtry_col_id,str(row_to_add_row_id));
            ws[usdtry_cell] = usdtry

        except:
            logger.error("IV) Error in updating FXRates sheet. Check the URL for sina fx in _get_sina_finance_fx")
            raise


        #PART V: Update TSIIndex
        try:
            #a) Get Data
            #login to thesteelindex to get a session
            session_requests = _get_login_session(type="steelindex")

            #get tsi price
            # prices_tsi = [
            #     {"cell_col_id":"B","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=1","name":"N.EU HRC"},
            #     {"cell_col_id":"C","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=9","name":"S.EU HRC"},
            #     {"cell_col_id":"D","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=93","name":"ASEAN HRC"},
            #     {"cell_col_id":"E","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=57","name":"Scrap"},
            #     {"cell_col_id":"F","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=52","name":"IO"},
            #     {"cell_col_id":"R","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=91","name":"Coking Coal"},
            # ]
            prices_tsi = [
                {"cell_col_id":"B","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=1","name":"N.EU HRC"},
                {"cell_col_id":"C","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=9","name":"S.EU HRC"},
                {"cell_col_id":"D","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=93","name":"ASEAN HRC"},
                {"cell_col_id":"E","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=57","name":"Scrap"},
                {"cell_col_id":"F","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=52","name":"IO"},
                {"cell_col_id":"G","value":None,"url":"https://www.thesteelindex.com/en/?cid=46&sid=91","name":"Coking Coal"},
            ]
            for price in prices_tsi:
                price['value'] = _get_thesteelindex_price(the_date=yesterday,session_requests=session_requests,url=price['url'],suppress_exception=suppress_exception)
                logger.info("{} price: {} (www.thesteelindex.com)".format(price['name'],str(price['value'])))

            #get feigang price
            # prices_feigang = [
            #     {"cell_col_id":"T","value":None,"url":"http://www.feigang.net/price.aspx?colid=120&area=&keyword=%E5%BC%A0%E5%AE%B6%E6%B8%AF","name":"Zhang Jiagang Scrap price"},
            # ]
            prices_feigang = [
                {"cell_col_id":"H","value":None,"url":"http://www.feigang.net/price.aspx?colid=120&area=&keyword=%E5%BC%A0%E5%AE%B6%E6%B8%AF","name":"Zhang Jiagang Scrap price"},
            ]
            for price in prices_feigang:
                price['value'] = _get_feigang_price(the_date=yesterday,url=price['url'],suppress_exception=suppress_exception)
                logger.info("{} price: {} (www.feigang.net)".format(price['name'],str(price['value'])))


            #b) Update TSIIndex excel file
            ws = wb['TSIIndex']

            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
            row_to_add_row_id = result['row_to_add_row_id']

            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            for price in prices_tsi:
                cell = '{}{}'.format(price["cell_col_id"],str(row_to_add_row_id));
                ws[cell] = price['value']
                ws[cell].number_format = ws['{}{}'.format(price["cell_col_id"],str(row_to_add_row_id-1))].number_format

            for price in prices_feigang:
                cell = '{}{}'.format(price["cell_col_id"],str(row_to_add_row_id));
                ws[cell] = float(price['value'])
                ws[cell].number_format = ws['{}{}'.format(price["cell_col_id"],str(row_to_add_row_id-1))].number_format

        except:
            logger.error("V) Error in updating TSIIndex sheet. Check the URL for sina fx in _get_sina_finance_fx and feigang in _get_feigang_price")
            raise


        #PART VI: Update DailyIndexes
        try:
            if False: #steelbb login got problem...
                #a) Get Data
                #login to steelbb to get a session
                session_requests = _get_login_session(type="steelbb")

                #get steelbb price [Not that sometime the data for yesterday would not be updated...]
                prices_steelbb = [
                    {"cell_col_id":"D","value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=431&period_type=5","name":"Blacksea Billet"},
                    {"cell_col_id":"E","value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=449&period_type=5","name":"Turkish Scrap"},
                    {"cell_col_id":"AC","value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=382&period_type=5","name":"Turkey Domestic Rebar(lira)"},
                    {"cell_col_id":"AE","value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=426&period_type=5","name":"Turkey Rebar Export"},
                ]
                for price in prices_steelbb:
                    price['value'] = _get_steelbb_price(the_date=yesterday,session_requests=session_requests,suppress_exception=suppress_exception,url=price['url'])['avg']
                    logger.info("{} price: {} (www.steelbb.com)".format(price['name'],str(price['value'])))



                #b) Update Daily Indexes excel file
                ws = wb['DailyIndexes']

                result = _find_row_to_add_by_date(ws=ws,date_col_idx=2,the_date=yesterday)
                row_to_add_row_id = result['row_to_add_row_id']

                _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="C",the_date=yesterday)

                for price in prices_steelbb:
                    cell = '{}{}'.format(price["cell_col_id"],str(row_to_add_row_id));
                    ws[cell] = price['value']
                    ws[cell].number_format = ws['{}{}'.format(price["cell_col_id"],str(row_to_add_row_id-1))].number_format


        except:
            logger.error("VI) Error in updating DailyIndexes sheet. Check the URL for sina fx in _get_steelbb_price.")
            raise


        #PART VII: Update AMMWeekly (Just insert new row)
        try:
            if False:
                ws = wb['AMMWeekly']
                previous_fri = yesterday #friday is 4
                while previous_fri.weekday()!= 4:
                    previous_fri = previous_fri - timedelta(days=1)

                result = _find_row_to_add_by_date(ws=ws,date_col_idx=2,the_date=previous_fri,previous_row_t_diff_day=-7)
                row_to_add_row_id = result['row_to_add_row_id']
                _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="C",the_date=previous_fri)
        except:
            logger.error("VII) Error in updating AMMWeekly.")
            raise





        #PART VIII: Update Dom-Exp Spread
        try:
            if False:
                ws = wb['Dom-Exp Spread']
                result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
                row_to_add_row_id = result['row_to_add_row_id']
                _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)
        except:
            logger.error("VIII) Error in updating Dom-Exp Spread. Check the URL for sina fx in _get_sina_finance_fx")
            raise


        #PART IX: Save to the network drive as tmp
        # wb.save(price_xlsx_network_tmp_path)

        #PART X: Save to local path...
        wb.save(price_xlsx_local_path)


        if(save_price_to_network):
            _copy(price_xlsx_network_tmp_path,price_xlsx_network_path)


        #PART X: Update SEA EAF.xlsx

    except:
        logger.error("update_prices_xlsx() failed for {}!".format(yesterday))
        raise


    logger.info("update_prices_xlsx() succeded for {}!".format(yesterday))


    pass





def update_sea_eaf(
                start_date=date.today(),
                the_date = None,
                temp_data_path = 'C:/tmp_data',
                sea_eaf_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/IFT SnD/SEA/SEA EAF_data.xlsx",
                sea_eaf_xlsx_local_relative_path = "/sea_eaf/SEA EAF_data.xlsx",
                copy_from_network = True,
                save_to_network = False,
                backup=True,
                suppress_exception=False
            ):
    '''
    S:\ETIA\DRY\FERROUS\Steel\IFT SnD\SEA\SEA EAF.xlsx
    REQUIRED SCRAPPING:
        _get_steelbb_price
    '''
    try:
        #PART I: PRE-PROCESS
        logger = logging.getLogger(__name__)

        if not the_date:
            today = start_date
            yesterday = get_biz_date(-1,today)
        else:
            yesterday = the_date
            today = get_biz_date(1,yesterday)
        day_before_yesterday = get_biz_date(-1,yesterday)
        today_str = get_date_str_frm_date(today)#get_date_str()
        yesterday_str = get_date_str_frm_date(yesterday)#get_date_str(-1)


        start_time = datetime.now()
        logger.info("Processing for today {} looking at the previous business day {} numbers".format(today.strftime("%Y-%m-%d (%a)"),yesterday.strftime("%Y-%m-%d (%a)")))



        #Copy sea eaf.xlsx to local.
        sea_eaf_xlsx_local_path = "{}{}".format(temp_data_path,sea_eaf_xlsx_local_relative_path)
        if copy_from_network:
            _copy(sea_eaf_xlsx_network_path,sea_eaf_xlsx_local_path)

        #Backup network copy first
        if backup:
            shutil.copy(sea_eaf_xlsx_local_path,sea_eaf_xlsx_local_path.replace(".xls","-{}.xls".format(today_str)))

        #load the excel file for writing
        wb = openpyxl.load_workbook(sea_eaf_xlsx_local_path)



        #PART II: Update Margin sheet
        try:
            #a) Get Data
            #login to steelbb to get a session
            session_requests = _get_login_session(type="steelbb")

            #get steelbb price [Not that sometime the data for yesterday would not be updated...]
            prices_steelbb = [
                {"cell_col_ids":{"B":"min","C":"max","D":"avg"},"value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=23&period_type=5","name":"Blacksea Billet"},
                {"cell_col_ids":{"E":"min","F":"max","G":"avg"},"value":None,"url":"https://www.steelbb.com/?PageID=93&series_id=36&period_type=5","name":"Blacksea Scrap"},
            ]
            for price in prices_steelbb:
                price['value'] = _get_steelbb_price(the_date=yesterday,session_requests=session_requests,url=price['url'],suppress_exception=suppress_exception)
                logger.info("{} price: min-{} max-{} avg={} (www.steelbb.com)".format(price['name'],str(price['value']['min']),str(price['value']['max']),str(price['value']['avg']) ))



            #b) Update Daily Indexes excel file
            ws = wb['Margin']

            previous_fri = yesterday #friday is 4
            while previous_fri.weekday()!= 4:
                previous_fri = previous_fri - timedelta(days=1)

            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday,previous_row_t_diff_day=-7)
            row_to_add_row_id = result['row_to_add_row_id']
            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            for price in prices_steelbb:
                for key, value in price["cell_col_ids"].items():
                    cell = '{}{}'.format(key,str(row_to_add_row_id));
                    ws[cell] = price['value'][value]
                    ws[cell].number_format = ws['{}{}'.format(key,str(row_to_add_row_id-1))].number_format
            pass
        except Exception, e:

            logger.error("II) Error in updating margin sheet for date {}. Error msg: {}".format(yesterday,str(e)))
            raise

        #PART IX: Save to the local drive as tmp
        wb.save(sea_eaf_xlsx_local_path)

        if(save_to_network):
            _copy(sea_eaf_xlsx_local_path,sea_eaf_xlsx_network_path)


    except Exception, e:
        logger.error("update_sea_eaf() failed for {}! Error msg: {}".format(yesterday,str(e)))
        raise


    logger.info("update_sea_eaf() succeeded for {}!".format(yesterday))











def load_mysteel_daily_price(logger=None):
    '''
    Load 日更数据
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    return load_data_from_mysteel_excel(
#                                start_date=ai.get_biz_date(-7), #by default load last 1 week data
                                start_date=None,
#                                end_date=ai.get_biz_date(1), #end_date is not included
                                end_date=None,
                                file_path=u"C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\MySteel Push email\\日更数据_new_mapping.xlsx",
                                overwrite_tags=True, #set to true if you want to overwrite the tags in case where the time series exists
                                overwrite_description=True,
                                start_row_excel_idx = 11, #by default we start reading from row 11 onwards
                                mapping={
                                            #key is the sheet name
                                            u"煤焦喷吹煤价格": {"series_name_row":2,"tags_row":3,"description_row":4,"start_data_col_excel":"C","end_data_col_excel":"J","ignore_cols":['D'],"start_row_excel_idx":12},

                                        },
                                logger=logger
                                )




def load_mysteel_rick_daily_price(logger=None):
    '''
    Load Rick's daily price into automan.
    '''
    if not logger:
        logger = logging.getLogger(__name__)

    return load_data_from_mysteel_excel(
#                                start_date=ai.get_biz_date(-7), #by default load last 1 week data
                                start_date=None,
#                                end_date=ai.get_biz_date(1), #end_date is not included
                                end_date=None,
                                file_path=u"C:\\Users\\j256377\\Desktop\\Metals\\Automan Project\\Rick's Daily Prices\\Daily prices_mapping.xlsx",
                                overwrite_tags=True, #set to true if you want to overwrite the tags in case where the time series exists
                                overwrite_description=True,
                                start_row_excel_idx = 11, #by default we start reading from row 11 onwards
                                mapping={
                                            #key is the sheet name
                                            # "Rebar prices": {"series_name_row":1,"tags_row":2,"description_row":3,"start_data_col_excel":"B","end_data_col_excel":"AB","ignore_cols":[]},
                                            # "PCI prices": {"series_name_row":1,"tags_row":2,"description_row":3,"start_data_col_excel":"B","end_data_col_excel":"U","ignore_cols":[]},
                                            # "Coke prices": {"series_name_row":1,"tags_row":2,"description_row":3,"start_data_col_excel":"B","end_data_col_excel":"U","ignore_cols":[]},
                                            # "HRC & Scrap prices": {"series_name_row":2,"tags_row":3,"description_row":4,"start_data_col_excel":"B","end_data_col_excel":"AC","ignore_cols":["E","L","Q"],"start_row_excel_idx":12},
                                            "Billet prices": {"series_name_row":2,"tags_row":3,"description_row":4,"start_data_col_excel":"B","end_data_col_excel":"C","ignore_cols":[],"start_row_excel_idx":12},

                                        },
                                logger=logger
                                )










def update_margin(
                start_date=date.today(),
                the_date = None,
                temp_data_path = 'C:/tmp_data',
                margin_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/John Dai/Margin_data.xlsx",
                margin_xlsx_local_relative_path = "/margin/Margin_data.xlsx",
                copy_from_network = True,
                save_to_network = False,
                backup=True,
                suppress_exception=False
            ):
    '''
    S:\ETIA\DRY\FERROUS\Steel\John Dai\Margin.xlsx
    REQUIRED SCRAPPING:
        _get_feigang_price
    '''

    try:
        #PART I: PRE-PROCESS
        logger = logging.getLogger(__name__)

        if not the_date:
            today = start_date
            yesterday = get_biz_date(-1,today)
        else:
            yesterday = the_date
            today = get_biz_date(1,yesterday)
        day_before_yesterday = get_biz_date(-1,yesterday)
        today_str = get_date_str_frm_date(today)#get_date_str()
        yesterday_str = get_date_str_frm_date(yesterday)#get_date_str(-1)


        start_time = datetime.now()
        logger.info("Processing for today {} looking at the previous business day {} numbers".format(today.strftime("%Y-%m-%d (%a)"),yesterday.strftime("%Y-%m-%d (%a)")))



        #Copy to local.
        margin_xlsx_local_path = "{}{}".format(temp_data_path,margin_xlsx_local_relative_path)
        if copy_from_network:
            _copy(margin_xlsx_network_path,margin_xlsx_local_path)

        #Backup network copy first
        if backup:
            shutil.copy(margin_xlsx_local_path,margin_xlsx_local_path.replace(".xlsx","-{}.xlsx".format(today_str)))

        #load the excel file for writing
        wb = openpyxl.load_workbook(margin_xlsx_local_path)



        #PART II: Update Raw Material Cost sheet
        try:
            #a) Get Data
            #get feigang price
            prices_feigang = [
                {"cell_col_id":"B","value":None,"url":"http://www.feigang.net/price.aspx?colid=120&area=&keyword=%E5%BC%A0%E5%AE%B6%E6%B8%AF","name":"Zhang Jiagang Scrap price"},
            ]
            for price in prices_feigang:
                price['value'] = _get_feigang_price(the_date=yesterday,url=price['url'],suppress_exception=suppress_exception)
                logger.info("{} price: {} (www.feigang.net)".format(price['name'],str(price['value'])))


            #b) Update sheet Raw Material Cost
            ws = wb['Raw Material Cost']

            result = _find_row_to_add_by_date(ws=ws,date_col_idx=0,the_date=yesterday)
            row_to_add_row_id = result['row_to_add_row_id']
            _update_date_col_in_ws(row_to_add_row_id,ws,date_col_id="A",the_date=yesterday)

            for price in prices_feigang:
                cell = '{}{}'.format(price["cell_col_id"],str(row_to_add_row_id));
                ws[cell] = float(price['value'])
                ws[cell].number_format = ws['{}{}'.format(price["cell_col_id"],str(row_to_add_row_id-1))].number_format



        except Exception, e:
            error_message = "II) Error in updating Raw Material Cost sheet for date {}. Error msg: {}".format(yesterday,str(e))
            logger.error(error_message)
            raise

        #PART IX: Save to the local drive as tmp
        wb.save(margin_xlsx_local_path)

        if(save_to_network):
            _copy(margin_xlsx_local_path,margin_xlsx_network_path)


    except Exception, e:
        logger.error("update_margin() failed for {}! Error msg: {}".format(yesterday,str(e)))
        raise


    logger.info("update_margin() succeeded for {}!".format(yesterday))

'''
template for new task
def update_new_task(
                start_date=date.today(),
                the_date = None,
                temp_data_path = 'C:/tmp_data',
                new_task_xlsx_network_path = "//sgsing022m/sing8data/ETIA/DRY/FERROUS/Steel/IFT SnD/SEA/SEA EAF_data.xlsx",
                new_task_xlsx_local_relative_path = "/new_task/SEA EAF_data.xls",
                copy_from_network = True,
                save_to_network = False,
                backup=True
            ):
    try:
        #PART I: PRE-PROCESS
        logger = logging.getLogger(__name__)

        if not the_date:
            today = start_date
            yesterday = get_biz_date(-1,today)
        else:
            yesterday = the_date
            today = get_biz_date(1,yesterday)
        day_before_yesterday = get_biz_date(-1,yesterday)
        today_str = get_date_str_frm_date(today)#get_date_str()
        yesterday_str = get_date_str_frm_date(yesterday)#get_date_str(-1)


        start_time = datetime.now()
        logger.info("Processing for today {} looking at the previous business day {} numbers".format(today.strftime("%Y-%m-%d (%a)"),yesterday.strftime("%Y-%m-%d (%a)")))



        #Copy to local.
        new_task_xlsx_local_path = "{}{}".format(temp_data_path,new_task_xlsx_local_relative_path)
        if copy_from_network:
            _copy(new_task_xlsx_network_path,new_task_xlsx_local_path)

        #Backup network copy first
        if backup:
            shutil.copy(new_task_xlsx_local_path,new_task_xlsx_local_path.replace(".xlsx","-{}.xlsx".format(today_str)))

        #load the excel file for writing
        wb = openpyxl.load_workbook(new_task_xlsx_local_path)



        #PART II: Update X sheet
        try:
            #a) Get Data
            pass


            #b) Update Daily Indexes excel file
            pass
        except Exception, e:

            logger.error("II) Error in updating X sheet for date {}. Error msg: {}".format(yesterday,str(e)))
            raise

        #PART IX: Save to the local drive as tmp
        wb.save(new_task_xlsx_local_path)

        if(save_to_network):
            _copy(new_task_xlsx_local_path,new_task_xlsx_network_path)


    except Exception, e:
        logger.error("update_new_task() failed for {}! Error msg: {}".format(yesterday,str(e)))
        raise


    logger.info("update_new_task() succeeded for {}!".format(yesterday))

'''
