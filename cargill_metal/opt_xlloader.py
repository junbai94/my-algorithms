# -*- coding: utf-8 -*-
"""
Created on Tue Oct 03 16:16:59 2017

@author: Junbai

This script is for uploading option trade book into database

To convert a string representation of dictionary to dictionary
>>> import json
>>> json.loads(json_string)
"""

import sqlite3 
import pandas as pd
from Data import Data
from openpyxl import load_workbook
import json
import xlwings as xw
import datetime

DEAL_DATABASE_PATH = "C:\\Users\\j291414\\Desktop\\deal_data.db"
DATABASE_PATH = "C:\\Users\\j291414\\Desktop\\market_data.db"
XL_PATH = "S:\\FERTRADE\\Bloomberg\\Trading Books\\BOF_Slave_New.xlsm"
VOL_PATH = "S:\\FERTRADE\\Bloomberg\\Trading Books\\dailyVolSnapshot.csv"
MACRO_PATH = "C:\\Users\\j291414\\Desktop\\macro.xlsm"
DEST_VOL_PATH = "C:\\Users\\j291414\\Desktop\\dailyVolSnapshot.csv"


def read_row(row):
    ret = dict()
    ret['strategy'] = row[0].value
    ret['book'] = 'BOF'
    ret['product'] = row[1].value
    ret['position'] = row[2].value
    ret['enter_date'] = str(row[5].value)
    ret['ccy'] = row[6].value
    ret['otype'] = row[8].value
    ret['strike'] = row[9].value
    ret['external_src'] = row[10].value
    ret['cpty'] = row[11].value
    ret['day1_comments'] = json.dumps({'premium':row[12].value, 'enter_vol':row[13].value}) 
    ret['status'] = 2
    ret['start'] = str(row[22].value)
    ret['end'] = str(row[23].value)
    return ret

def instrument(ret):
    spec = dict()
    spec['product'] = ret['product']
    spec['inst_type'] = 'ComMthAsian'
    spec['ccy'] = ret['ccy']
    spec['otype'] = ret['otype']
    spec['strike'] = ret['strike']
    spec['start'] = ret['start']
    spec['end'] = ret['end']
    return json.dumps([[spec, ret['position']],])

def to_db(ret, instruments, conn):
    c = conn.cursor()
    sql = "insert into deals (positions, enter_date, last_date, strategy, book, external_src, cpty, day1_comments, status) \
    values (?, ?, ?, ?, ?, ?, ?, ?, ?)"
    c.execute(sql, (instruments, ret['enter_date'], ret['end'],
                    ret['strategy'], ret['book'], ret['external_src'], ret['cpty'],
                    ret['day1_comments'], ret['status']))
    
    

    
'''
+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++'''

def main():     # update option deal data
    conn = sqlite3.connect(DEAL_DATABASE_PATH)
    wb = load_workbook(XL_PATH, data_only=True)
    ws = wb['Ins_AsianOption']
    
    conn.execute("delete from deals")
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=25):
        ret = read_row(row)
        if ret['position'] == None:
            break
        instruments = instrument(ret)
        to_db(ret, instruments, conn)
        
    conn.commit()
    conn.close()
    
def main2():    # update broker vol quote
    exch = 'SGX'
    prodCode = 'fef'
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    xb = xw.Book(MACRO_PATH)
    func = xb.macro('DailyVolSnapshot')
    func(VOL_PATH, DEST_VOL_PATH)
    xb.close()
    df = pd.read_csv(DEST_VOL_PATH)
    c.execute("delete from cmvol_daily")
    for i in range(len(df)):
        sql = "insert into cmvol_daily values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        c.execute(sql, (df.iloc[i,0], str(df.iloc[i,1]), df.iloc[i,2], df.iloc[i,3],
                        df.iloc[i,4], df.iloc[i,5], df.iloc[i,6], df.iloc[i,7],
                        exch, prodCode))
    
    conn.commit()
    conn.close()

if __name__ == '__main__':
    main()
    main2()


